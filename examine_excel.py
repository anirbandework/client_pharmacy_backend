#!/usr/bin/env python3

import sys
import os
sys.path.append('/Users/anirbande/Desktop/client backend')

from openpyxl import load_workbook
import pandas as pd

def examine_excel_file():
    file_path = '/Users/anirbande/Desktop/client backend/modules/daily_records/GMTR0003 business records .xlsx'
    
    try:
        # Load workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        print(f"Sheet name: {sheet.title}")
        print(f"Max row: {sheet.max_row}")
        print(f"Max column: {sheet.max_column}")
        print()
        
        # Print first 10 rows to see structure
        print("First 10 rows:")
        for row_num in range(1, min(11, sheet.max_row + 1)):
            row_data = []
            for col_num in range(1, min(18, sheet.max_column + 1)):  # Check first 17 columns
                cell_value = sheet.cell(row=row_num, column=col_num).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            print(f"Row {row_num}: {row_data}")
        
        print("\nData starting from row 6 (where import should start):")
        for row_num in range(6, min(16, sheet.max_row + 1)):  # Check 10 rows from row 6
            row_data = []
            for col_num in range(1, 18):  # Columns A to Q
                cell_value = sheet.cell(row=row_num, column=col_num).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            print(f"Row {row_num}: {row_data}")
            
        # Check specific columns that should have data
        print("\nChecking specific columns for actual data:")
        print("Column A (Date), Column 3 (Cash Balance), Column 7 (Actual Cash):")
        for row_num in range(6, min(16, sheet.max_row + 1)):
            date_val = sheet.cell(row=row_num, column=1).value
            cash_balance = sheet.cell(row=row_num, column=3).value
            actual_cash = sheet.cell(row=row_num, column=7).value
            print(f"Row {row_num}: Date={date_val}, Cash Balance={cash_balance}, Actual Cash={actual_cash}")
            
    except Exception as e:
        print(f"Error reading Excel file: {e}")

if __name__ == "__main__":
    examine_excel_file()