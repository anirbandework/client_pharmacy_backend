from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database.database import get_db
from datetime import datetime, date, timedelta
from typing import Optional, List
from . import schemas, models, services
from modules.auth.dependencies import get_current_user as get_user_dict
from modules.auth.models import Staff, Shop
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

def get_current_user(user_dict: dict = Depends(get_user_dict), db: Session = Depends(get_db)) -> tuple[Staff, int]:
    """Extract staff user from auth dict and resolve shop_id from shop_code"""
    if user_dict["token_data"].user_type != "staff":
        raise HTTPException(status_code=403, detail="Staff access required")
    
    staff = user_dict["user"]
    shop_code = user_dict["token_data"].shop_code
    
    if not shop_code:
        raise HTTPException(status_code=400, detail="Shop code not found in token")
    
    # Resolve shop_id from shop_code using organization_id
    shop = db.query(Shop).filter(
        Shop.shop_code == shop_code,
        Shop.organization_id == staff.shop.organization_id
    ).first()
    if not shop:
        raise HTTPException(status_code=404, detail=f"Shop not found with code: {shop_code}")
    
    return staff, shop.id

router = APIRouter()

# MEDICINE SEARCH

@router.get("/search-medicines", response_model=List[schemas.MedicineSearchResult])
def search_medicines(
    q: str = Query(..., min_length=2, description="Search term"),
    limit: int = Query(20, le=100),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Search medicines by name, generic name, brand, or batch number with location info"""
    staff, shop_id = current_user
    results = services.BillingService.search_medicines(db, shop_id, q, limit)
    return results

# BILL MANAGEMENT

@router.post("/bills", response_model=schemas.BillResponse)
def create_bill(
    bill_data: schemas.BillCreate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Create new bill and update stock automatically"""
    staff, shop_id = current_user
    
    try:
        bill = services.BillingService.create_bill(
            db,
            shop_id,
            staff.id,
            staff.name,
            bill_data.model_dump(exclude={'items'}),
            [item.model_dump() for item in bill_data.items]
        )
        return bill
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/bills", response_model=List[schemas.BillResponse])
def get_bills(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    customer_phone: Optional[str] = None,
    payment_method: Optional[models.PaymentMethod] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get bills with filters"""
    staff, shop_id = current_user
    query = db.query(models.Bill).filter(models.Bill.shop_id == shop_id)
    
    if start_date:
        query = query.filter(models.Bill.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.filter(models.Bill.created_at <= datetime.combine(end_date, datetime.max.time()))
    if customer_phone:
        query = query.filter(models.Bill.customer_phone == customer_phone)
    if payment_method:
        query = query.filter(models.Bill.payment_method == payment_method)
    
    return query.order_by(models.Bill.created_at.desc()).offset(skip).limit(limit).all()

@router.get("/bills/{bill_id}", response_model=schemas.BillResponse)
def get_bill(
    bill_id: int,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get specific bill details"""
    staff, shop_id = current_user
    bill = db.query(models.Bill).filter(
        models.Bill.id == bill_id,
        models.Bill.shop_id == shop_id
    ).first()
    
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    
    return bill

@router.get("/bills/number/{bill_number}", response_model=schemas.BillResponse)
def get_bill_by_number(
    bill_number: str,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get bill by bill number"""
    staff, shop_id = current_user
    bill = db.query(models.Bill).filter(
        models.Bill.bill_number == bill_number,
        models.Bill.shop_id == shop_id
    ).first()
    
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    
    return bill

@router.delete("/bills/{bill_id}")
def delete_bill(
    bill_id: int,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Delete bill (admin only - restores stock)"""
    staff, shop_id = current_user
    
    if not staff.can_manage_staff:  # Only managers can delete bills
        raise HTTPException(status_code=403, detail="Permission denied")
    
    bill = db.query(models.Bill).filter(
        models.Bill.id == bill_id,
        models.Bill.shop_id == shop_id
    ).first()
    
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    
    # Restore stock quantities and recalculate discrepancy
    from modules.stock_audit.models import StockItem
    for item in bill.items:
        stock_item = db.query(StockItem).filter(
            StockItem.id == item.stock_item_id,
            StockItem.shop_id == shop_id
        ).first()
        if stock_item:
            stock_item.quantity_software += item.quantity
            if stock_item.quantity_physical is not None:
                stock_item.audit_discrepancy = stock_item.quantity_software - stock_item.quantity_physical
            stock_item.updated_at = datetime.utcnow()
    
    db.delete(bill)
    db.commit()
    return {"message": "Bill deleted and stock restored"}

# REPORTS

@router.get("/summary", response_model=schemas.BillSummary)
def get_billing_summary(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get billing summary for date range"""
    staff, shop_id = current_user
    return services.BillingService.get_bill_summary(db, shop_id, start_date, end_date)

@router.get("/top-selling")
def get_top_selling_items(
    limit: int = Query(10, le=50),
    days: int = Query(30, le=365),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get top selling items"""
    staff, shop_id = current_user
    return services.BillingService.get_top_selling_items(db, shop_id, limit, days)

@router.get("/customer-history/{customer_phone}")
def get_customer_history(
    customer_phone: str,
    limit: int = Query(50, le=200),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get customer purchase history"""
    staff, shop_id = current_user
    bills = db.query(models.Bill).filter(
        models.Bill.shop_id == shop_id,
        models.Bill.customer_phone == customer_phone
    ).order_by(models.Bill.created_at.desc()).limit(limit).all()
    
    total_spent = sum(b.total_amount for b in bills)
    
    return {
        "customer_phone": customer_phone,
        "total_bills": len(bills),
        "total_spent": float(total_spent),
        "bills": bills
    }

@router.get("/daily-sales")
def get_daily_sales(
    days: int = Query(7, le=90),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get daily sales for last N days"""
    staff, shop_id = current_user
    cutoff_date = datetime.utcnow() - timedelta(days=days)
    
    from sqlalchemy import func
    results = db.query(
        func.date(models.Bill.created_at).label('date'),
        func.count(models.Bill.id).label('bill_count'),
        func.sum(models.Bill.total_amount).label('total_sales')
    ).filter(
        models.Bill.shop_id == shop_id,
        models.Bill.created_at >= cutoff_date
    ).group_by(
        func.date(models.Bill.created_at)
    ).order_by(
        func.date(models.Bill.created_at).desc()
    ).all()
    
    return [
        {
            "date": str(r.date),
            "bill_count": r.bill_count,
            "total_sales": float(r.total_sales)
        }
        for r in results
    ]

# EXCEL EXPORT

@router.get("/export/bills")
def export_bills_excel(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Export bills to Excel"""
    staff, shop_id = current_user
    query = db.query(models.Bill).filter(models.Bill.shop_id == shop_id)
    
    if start_date:
        query = query.filter(models.Bill.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.filter(models.Bill.created_at <= datetime.combine(end_date, datetime.max.time()))
    
    bills = query.order_by(models.Bill.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Bills"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ["Bill Number", "Date", "Customer Name", "Customer Phone", "Doctor Name", "Payment Method", 
               "Subtotal", "Discount", "Tax", "Total Amount", "Amount Paid", "Change", "Staff Name"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for row, bill in enumerate(bills, 2):
        ws.cell(row=row, column=1, value=bill.bill_number)
        ws.cell(row=row, column=2, value=bill.created_at.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=3, value=bill.customer_name or "")
        ws.cell(row=row, column=4, value=bill.customer_phone or "")
        ws.cell(row=row, column=5, value=bill.doctor_name or "")
        ws.cell(row=row, column=6, value=bill.payment_method.value)
        ws.cell(row=row, column=7, value=bill.subtotal)
        ws.cell(row=row, column=8, value=bill.discount_amount)
        ws.cell(row=row, column=9, value=bill.tax_amount)
        ws.cell(row=row, column=10, value=bill.total_amount)
        ws.cell(row=row, column=11, value=bill.amount_paid)
        ws.cell(row=row, column=12, value=bill.change_returned)
        ws.cell(row=row, column=13, value=bill.staff_name)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"bills_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
