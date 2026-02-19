from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database.database import get_db
from datetime import date
from typing import List
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from . import daily_records_schemas as schemas
from .daily_records_service import DailyRecordsService
from .daily_records_models import DailyRecord, DailyExpense
from modules.auth.dependencies import get_current_user as get_user_dict
from modules.auth.models import Staff, Shop

def get_current_user(user_dict: dict = Depends(get_user_dict), db: Session = Depends(get_db)) -> tuple[Staff, int]:
    """Extract staff user from auth dict and resolve shop_id from shop_code"""
    if user_dict["token_data"].user_type != "staff":
        raise HTTPException(status_code=403, detail="Staff access required")
    
    staff = user_dict["user"]
    shop_code = user_dict["token_data"].shop_code
    
    if not shop_code:
        raise HTTPException(status_code=400, detail="Shop code not found in token")
    
    shop = db.query(Shop).filter(
        Shop.shop_code == shop_code,
        Shop.organization_id == staff.shop.organization_id
    ).first()
    if not shop:
        raise HTTPException(status_code=404, detail=f"Shop not found with code: {shop_code}")
    
    return staff, shop.id

router = APIRouter()

@router.get("/daily-records/{record_date}", response_model=schemas.DailyRecordResponse)
def get_daily_record(
    record_date: date,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get daily record for specific date"""
    staff, shop_id = current_user
    
    record = db.query(DailyRecord).filter(
        DailyRecord.shop_id == shop_id,
        DailyRecord.record_date == record_date
    ).first()
    
    if not record:
        # Create empty record with auto-calculated figures
        record = DailyRecordsService.update_daily_record(
            db, shop_id, record_date, staff.id, staff.name, {}
        )
    
    return DailyRecordsService.get_daily_record_with_calculations(db, record)

@router.post("/daily-records", response_model=schemas.DailyRecordResponse)
def create_or_update_daily_record(
    data: schemas.DailyRecordCreate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Create or update daily record"""
    staff, shop_id = current_user
    
    record = DailyRecordsService.update_daily_record(
        db,
        shop_id,
        data.record_date,
        staff.id,
        staff.name,
        data.model_dump(exclude={'expenses'})
    )
    
    # Add expenses
    for expense_data in data.expenses:
        DailyRecordsService.add_expense(
            db,
            shop_id,
            record.id,
            expense_data.model_dump()
        )
    
    db.refresh(record)
    return DailyRecordsService.get_daily_record_with_calculations(db, record)

@router.put("/daily-records/{record_date}", response_model=schemas.DailyRecordResponse)
def update_daily_record(
    record_date: date,
    data: schemas.DailyRecordUpdate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Update daily record"""
    staff, shop_id = current_user
    
    record = DailyRecordsService.update_daily_record(
        db,
        shop_id,
        record_date,
        staff.id,
        staff.name,
        data.model_dump(exclude_unset=True)
    )
    
    return DailyRecordsService.get_daily_record_with_calculations(db, record)

@router.post("/daily-records/{record_date}/expenses", response_model=schemas.DailyExpense)
def add_expense(
    record_date: date,
    expense: schemas.DailyExpenseCreate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Add expense to daily record"""
    staff, shop_id = current_user
    
    record = db.query(DailyRecord).filter(
        DailyRecord.shop_id == shop_id,
        DailyRecord.record_date == record_date
    ).first()
    
    if not record:
        record = DailyRecordsService.get_or_create_daily_record(
            db, shop_id, record_date, staff.id, staff.name
        )
    
    return DailyRecordsService.add_expense(
        db,
        shop_id,
        record.id,
        expense.model_dump(),
        staff.id,
        staff.name
    )

@router.delete("/daily-records/expenses/{expense_id}")
def delete_expense(
    expense_id: int,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Delete expense"""
    staff, shop_id = current_user
    
    expense = db.query(DailyExpense).filter(
        DailyExpense.id == expense_id,
        DailyExpense.shop_id == shop_id
    ).first()
    
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    daily_record_id = expense.daily_record_id
    db.delete(expense)
    
    # Update total expenses
    record = db.query(DailyRecord).filter(DailyRecord.id == daily_record_id).first()
    if record:
        from sqlalchemy import func
        total = db.query(func.sum(DailyExpense.amount)).filter(
            DailyExpense.daily_record_id == daily_record_id
        ).scalar() or 0.0
        record.total_expenses = float(total)
    
    db.commit()
    return {"message": "Expense deleted"}

@router.get("/daily-records", response_model=List[schemas.DailyRecordResponse])
def get_daily_records(
    start_date: date = Query(...),
    end_date: date = Query(...),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get daily records for date range"""
    staff, shop_id = current_user
    
    records = db.query(DailyRecord).filter(
        DailyRecord.shop_id == shop_id,
        DailyRecord.record_date >= start_date,
        DailyRecord.record_date <= end_date
    ).order_by(DailyRecord.record_date.desc()).all()
    
    return [DailyRecordsService.get_daily_record_with_calculations(db, r) for r in records]

@router.get("/daily-records/export/excel")
def export_daily_records_excel(
    start_date: date = Query(...),
    end_date: date = Query(...),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Export daily records to Excel"""
    staff, shop_id = current_user
    
    if start_date > end_date:
        raise HTTPException(status_code=400, detail="Start date cannot be after end date")
    
    records = db.query(DailyRecord).filter(
        DailyRecord.shop_id == shop_id,
        DailyRecord.record_date >= start_date,
        DailyRecord.record_date <= end_date
    ).order_by(DailyRecord.record_date).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Records"
    
    # Headers for Daily Records
    headers = [
        "Date", "Day", "No. of Bills", "Average Bill", "Software Sales", "Cash Sales", 
        "Depositable Amount", "Small Denomination", "Online Sales", "Total Sales", 
        "Unbilled Amount", "Recorded Sales", "Difference", "Cash Deposited", 
        "Cash Reserve", "Total Expenses", "Expense Details"
    ]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data for Daily Records
    for row_idx, record in enumerate(records, 2):
        calc = DailyRecordsService.get_daily_record_with_calculations(db, record)
        record_date = calc['record_date']
        day_name = record_date.strftime('%A') if hasattr(record_date, 'strftime') else ''
        
        # Build expense details string
        expense_details = ""
        if calc.get('expenses'):
            expense_lines = []
            for exp in calc['expenses']:
                expense_lines.append(f"{exp.expense_category}: ₹{exp.amount:.2f} ({exp.description or 'N/A'}) - {exp.staff_name or 'Unknown'}")
            expense_details = "; ".join(expense_lines)
        
        ws.cell(row=row_idx, column=1, value=str(record_date))
        ws.cell(row=row_idx, column=2, value=day_name)
        ws.cell(row=row_idx, column=3, value=calc['no_of_bills'])
        ws.cell(row=row_idx, column=4, value=round(calc['average_bill'], 2))
        ws.cell(row=row_idx, column=5, value=round(calc['software_sales'], 2))
        ws.cell(row=row_idx, column=6, value=round(calc['cash_sales'], 2))
        ws.cell(row=row_idx, column=7, value=round(calc['depositable_amount'], 2))
        ws.cell(row=row_idx, column=8, value=round(calc['small_denomination'], 2))
        ws.cell(row=row_idx, column=9, value=round(calc['online_sales'], 2))
        ws.cell(row=row_idx, column=10, value=round(calc['total_sales'], 2))
        ws.cell(row=row_idx, column=11, value=round(calc['unbilled_amount'], 2))
        ws.cell(row=row_idx, column=12, value=round(calc['recorded_sales'], 2))
        ws.cell(row=row_idx, column=13, value=round(calc['difference'], 2))
        ws.cell(row=row_idx, column=14, value=round(calc['actual_cash_deposited'], 2))
        ws.cell(row=row_idx, column=15, value=round(calc['cash_reserve'], 2))
        ws.cell(row=row_idx, column=16, value=round(calc['total_expenses'], 2))
        ws.cell(row=row_idx, column=17, value=expense_details)
    
    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Create Expenses sheet
    ws_expenses = wb.create_sheet(title="Expenses")
    expense_headers = ["Date", "Day", "Category", "Amount", "Description", "Staff", "Time"]
    
    for col, header in enumerate(expense_headers, 1):
        cell = ws_expenses.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add expense data
    expense_row = 2
    for record in records:
        calc = DailyRecordsService.get_daily_record_with_calculations(db, record)
        record_date = calc['record_date']
        day_name = record_date.strftime('%A') if hasattr(record_date, 'strftime') else ''
        
        for expense in calc.get('expenses', []):
            ws_expenses.cell(row=expense_row, column=1, value=str(record_date))
            ws_expenses.cell(row=expense_row, column=2, value=day_name)
            ws_expenses.cell(row=expense_row, column=3, value=expense.expense_category)
            ws_expenses.cell(row=expense_row, column=4, value=round(expense.amount, 2))
            ws_expenses.cell(row=expense_row, column=5, value=expense.description or "")
            ws_expenses.cell(row=expense_row, column=6, value=expense.staff_name or "")
            ws_expenses.cell(row=expense_row, column=7, value=str(expense.created_at))
            expense_row += 1
    
    # Auto-size expense columns
    for col in ws_expenses.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_expenses.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"daily_records_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
