from .models import RecordModification, DailyRecord
from .schemas import ExcelImportResponse
from sqlalchemy.orm import Session
from fastapi import UploadFile
from datetime import datetime, date
import io

def send_whatsapp_alert(message: str):
    """Send WhatsApp alert for cash differences"""
    try:
        import requests
        from app.core.config import settings
        
        if hasattr(settings, 'WHATSAPP_API_URL') and hasattr(settings, 'WHATSAPP_TOKEN'):
            headers = {
                'Authorization': f'Bearer {settings.WHATSAPP_TOKEN}',
                'Content-Type': 'application/json'
            }
            data = {
                'messaging_product': 'whatsapp',
                'to': settings.WHATSAPP_PHONE_NUMBER,
                'type': 'text',
                'text': {'body': message}
            }
            requests.post(settings.WHATSAPP_API_URL, json=data, headers=headers)
    except Exception as e:
        print(f"WhatsApp alert failed: {e}")

def check_cash_difference_alert(sales_difference: float, record_date, total_sales: float):
    """Check if sales difference exceeds threshold and send alert"""
    try:
        from app.core.config import settings
        threshold = getattr(settings, 'cash_difference_threshold', 100)
        if abs(sales_difference) > threshold:
            send_whatsapp_alert(
                f"⚠️ Cash Variance Alert\n\nDate: {record_date}\nDifference: ₹{sales_difference:.2f}\nTotal Sales: ₹{total_sales:.2f}\n\nPlease investigate."
            )
            return True
    except Exception as e:
        print(f"Alert error: {e}")
    return False

def track_modifications(db: Session, record, update_data: dict, record_id: int, modified_by: str = None):
    """Track all modifications to a record"""
    modifications = []
    
    for field, new_value in update_data.items():
        if new_value is not None:
            old_value = getattr(record, field)
            if old_value != new_value:
                modification = RecordModification(
                    daily_record_id=record_id,
                    field_name=field,
                    old_value=str(old_value) if old_value is not None else 'None',
                    new_value=str(new_value),
                    modified_by=modified_by or 'system'
                )
                modifications.append(modification)
    
    for mod in modifications:
        db.add(mod)
    
    return modifications

def calculate_fields(record_data: dict):
    """Calculate derived fields based on Excel formulas"""
    # Extract values with defaults
    actual_cash = float(record_data.get('actual_cash') or 0)
    cash_reserve = float(record_data.get('cash_reserve') or 0)
    expense_amount = float(record_data.get('expense_amount') or 0)
    online = float(record_data.get('online_sales') or 0)
    unbilled = float(record_data.get('unbilled_sales') or 0)
    software_fig = float(record_data.get('software_figure') or 0)
    no_of_bills = int(record_data.get('no_of_bills') or 0)
    
    # Total Cash = Actual Cash + Cash Reserve + Expense Amount
    total_cash = actual_cash + cash_reserve + expense_amount
    
    # Total Sales = Actual Cash + Online
    total_sales = actual_cash + online
    
    # Recorded Sales = Unbilled Sales + Software Figure
    recorded_sales = unbilled + software_fig
    
    # Sales Difference = Total Sales - Recorded Sales
    sales_difference = total_sales - recorded_sales
    
    # Average Bill = Recorded Sales / Number of Bills
    average_bill = recorded_sales / no_of_bills if no_of_bills > 0 else 0
    
    return {
        'total_cash': round(total_cash, 2),
        'total_sales': round(total_sales, 2),
        'recorded_sales': round(recorded_sales, 2),
        'sales_difference': round(sales_difference, 2),
        'average_bill': round(average_bill, 2)
    }

async def import_excel_data(file: UploadFile, db: Session) -> ExcelImportResponse:
    """Import daily record data from Excel file matching GMTR0003 format"""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import range_boundaries
        
        content = await file.read()
        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook.active
        
        records_imported = 0
        errors = []
        
        # Start from row 6 (first data row after headers)
        for row_num in range(6, sheet.max_row + 1):
            try:
                # Extract date from column A
                date_val = sheet.cell(row=row_num, column=1).value
                if not date_val:
                    continue
                    
                # Parse date (handle Excel serial numbers)
                if isinstance(date_val, datetime):
                    record_date = date_val.date()
                elif isinstance(date_val, date):
                    record_date = date_val
                elif isinstance(date_val, (int, float)):
                    # Excel serial number (days since 1900-01-01)
                    from datetime import timedelta
                    excel_epoch = datetime(1899, 12, 30)  # Excel's epoch
                    record_date = (excel_epoch + timedelta(days=date_val)).date()
                else:
                    record_date = datetime.strptime(str(date_val), '%Y-%m-%d').date()
                
                # Helper function to get cell value (handles formulas)
                def get_cell_value(row, col, default=0):
                    cell = sheet.cell(row=row, column=col)
                    value = cell.value
                    if value is None:
                        return default
                    
                    # If it's a string that starts with '=', it's a formula
                    if isinstance(value, str) and value.startswith('='):
                        # Try to extract numbers from simple formulas
                        formula = value[1:]  # Remove the '=' sign
                        
                        # Handle simple addition formulas like "9500+2000"
                        if '+' in formula and all(c.isdigit() or c in '+- .' for c in formula):
                            try:
                                return eval(formula)
                            except:
                                pass
                        
                        # Handle sum formulas like "sum(G6+N6+Q6)+1203"
                        if 'sum(' in formula.lower():
                            # For now, return default for complex sum formulas
                            return default
                            
                        # Try to extract the first number from the formula
                        import re
                        numbers = re.findall(r'\d+(?:\.\d+)?', formula)
                        if numbers:
                            return float(numbers[0])
                        
                        return default
                    
                    return value
                
                # Extract data from Excel columns (matching GMTR0003 format)
                record_data = {
                    'date': record_date,
                    'day': str(get_cell_value(row_num, 2, '')),
                    'cash_balance': float(get_cell_value(row_num, 3, 1000)),
                    'no_of_bills': int(get_cell_value(row_num, 5, 0)),
                    'actual_cash': float(get_cell_value(row_num, 7, 0)),
                    'online_sales': float(get_cell_value(row_num, 8, 0)),
                    'unbilled_sales': float(get_cell_value(row_num, 10, 0)),
                    'software_figure': float(get_cell_value(row_num, 11, 0)),
                    'cash_reserve': float(get_cell_value(row_num, 14, 0)),
                    'reserve_comments': str(get_cell_value(row_num, 15, '')),
                    'expense_amount': float(get_cell_value(row_num, 17, 0)),
                    'notes': str(get_cell_value(row_num, 15, '')),
                    'created_by': 'excel_import',
                    'shop_id': None  # Set to None since it's nullable
                }
                
                # Calculate derived fields
                calculated = calculate_fields(record_data)
                record_data.update(calculated)
                
                # Check if record already exists
                existing = db.query(DailyRecord).filter(DailyRecord.date == record_date).first()
                if existing:
                    # Update existing record instead of skipping
                    for field, value in record_data.items():
                        if field != 'date':  # Don't update the date
                            setattr(existing, field, value)
                    
                    # Calculate derived fields for existing record
                    calculated = calculate_fields(record_data)
                    for field, value in calculated.items():
                        setattr(existing, field, value)
                    
                    existing.modified_at = datetime.utcnow()
                    existing.modified_by = 'excel_import_update'
                    records_imported += 1
                else:
                    # Create new record
                    db_record = DailyRecord(**record_data)
                    db.add(db_record)
                    records_imported += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        db.commit()
        
        return ExcelImportResponse(
            success=True,
            records_imported=records_imported,
            errors=errors,
            message=f"Successfully imported {records_imported} records" + (f" with {len(errors)} errors" if errors else "")
        )
        
    except Exception as e:
        db.rollback()
        return ExcelImportResponse(
            success=False,
            records_imported=0,
            errors=[str(e)],
            message="Failed to import Excel file"
        )

def export_to_excel(records: list, year: int, month: int) -> bytes:
    """Export records to Excel in GMTR0003 format"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Title row
        ws.merge_cells('A1:Q1')
        ws['A1'] = f"GMTR0003 business record - employee update sheet"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers (Row 3)
        headers_row1 = ['Date', 'Day', 'Cash\nbalance', 'Average bill', 'No.\nof bills', 
                        'Cash', None, 'Online', 'Total sales', 'Sales', None, None, 
                        'Difference', 'Cash Reserve', 'reserve balance', None, 'Expense record']
        headers_row2 = [None, None, None, None, None, 'Total Cash', 'Actual Cash', None, None,
                        'Unbilled', 'Software fig.', 'Recorded sales', None, None, 'Comments', None, 'Amount']
        
        for col, header in enumerate(headers_row1, start=1):
            if header:
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col, header in enumerate(headers_row2, start=1):
            if header:
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows starting from row 6
        for row_idx, record in enumerate(records, start=6):
            ws.cell(row=row_idx, column=1, value=record.date)
            ws.cell(row=row_idx, column=2, value=record.day)
            ws.cell(row=row_idx, column=3, value=record.cash_balance)
            ws.cell(row=row_idx, column=4, value=record.average_bill)
            ws.cell(row=row_idx, column=5, value=record.no_of_bills)
            ws.cell(row=row_idx, column=6, value=record.total_cash)
            ws.cell(row=row_idx, column=7, value=record.actual_cash)
            ws.cell(row=row_idx, column=8, value=record.online_sales)
            ws.cell(row=row_idx, column=9, value=record.total_sales)
            ws.cell(row=row_idx, column=10, value=record.unbilled_sales)
            ws.cell(row=row_idx, column=11, value=record.software_figure)
            ws.cell(row=row_idx, column=12, value=record.recorded_sales)
            ws.cell(row=row_idx, column=13, value=record.sales_difference)
            ws.cell(row=row_idx, column=14, value=record.cash_reserve)
            ws.cell(row=row_idx, column=15, value=record.reserve_comments)
            ws.cell(row=row_idx, column=17, value=record.expense_amount)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        raise Exception(f"Failed to export to Excel: {str(e)}")