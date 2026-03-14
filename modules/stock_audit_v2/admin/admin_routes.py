from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database.database import get_db
from datetime import datetime, date, timedelta
from typing import Optional
import math
from .. import schemas, models
from .admin_analytics_service import StockAuditAnalytics
from .admin_ai_analytics_service import StockAuditAIAnalytics
from modules.auth.dependencies import get_current_admin
from modules.auth.models import Admin, Shop
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from app.utils.cache import dashboard_cache

router = APIRouter()

# ADMIN RACK AND SECTION VIEWS (Read-only)

@router.get("/racks", response_model=list[schemas.StoreRack])
def get_admin_racks(
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Get all store racks for admin (org-scoped)"""
    return (
        db.query(models.StockRack)
        .join(Shop, models.StockRack.shop_id == Shop.id)
        .filter(Shop.organization_id == admin.organization_id)
        .all()
    )

@router.get("/sections", response_model=list[schemas.StoreSection])
def get_admin_sections(
    rack_id: Optional[int] = None,
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Get all sections for admin (org-scoped)"""
    query = (
        db.query(models.StockSection)
        .join(Shop, models.StockSection.shop_id == Shop.id)
        .filter(Shop.organization_id == admin.organization_id)
    )
    if rack_id:
        query = query.filter(models.StockSection.rack_id == rack_id)
    return query.all()

# ADMIN EXCEL UPLOAD MANAGEMENT

@router.get("/uploads")
def get_admin_excel_uploads(
    status: Optional[str] = None,
    search: Optional[str] = None,
    shop_id: Optional[int] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Get Excel uploads for admin (org-scoped, paginated, searchable)"""
    import math
    query = (
        db.query(models.ExcelUpload)
        .join(Shop, models.ExcelUpload.shop_id == Shop.id)
        .filter(Shop.organization_id == admin.organization_id)
    )

    if status:
        query = query.filter(models.ExcelUpload.status == status)
    if shop_id:
        query = query.filter(models.ExcelUpload.shop_id == shop_id)
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            (models.ExcelUpload.filename.ilike(search_term)) |
            (models.ExcelUpload.uploaded_by_staff_name.ilike(search_term))
        )

    total = query.count()
    uploads = query.order_by(models.ExcelUpload.uploaded_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

    result = []
    for upload in uploads:
        result.append({
            "id": upload.id,
            "filename": upload.filename,
            "uploaded_by": upload.uploaded_by_staff_name,
            "uploaded_at": upload.uploaded_at,
            "status": upload.status,
            "total_items": upload.total_items,
            "success_count": upload.success_count,
            "error_count": upload.error_count,
            "staff_verified": upload.staff_verified,
            "staff_verified_by": upload.staff_verified_by_name,
            "staff_verified_at": upload.staff_verified_at,
            "admin_verified": upload.admin_verified,
            "admin_verified_by": upload.admin_verified_by_name,
            "admin_verified_at": upload.admin_verified_at,
            "upload_notes": upload.upload_notes,
            "staff_notes": upload.staff_notes,
            "admin_notes": upload.admin_notes,
            "rejection_reason": upload.rejection_reason
        })

    return {
        "items": result,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": math.ceil(total / per_page) if total > 0 else 1
    }

@router.get("/uploads/{upload_id}/items")
def get_admin_upload_items(
    upload_id: int,
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Get items from specific upload (admin)"""
    upload = db.query(models.ExcelUpload).join(Shop, models.ExcelUpload.shop_id == Shop.id).filter(
        models.ExcelUpload.id == upload_id,
        Shop.organization_id == admin.organization_id
    ).first()
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    items = db.query(models.ExcelUploadItem).filter(
        models.ExcelUploadItem.upload_id == upload_id
    ).all()
    
    result = []
    for item in items:
        result.append({
            "id": item.id,
            "product_name": item.product_name,
            "composition": item.composition,
            "manufacturer": item.manufacturer,
            "hsn_code": item.hsn_code,
            "batch_number": item.batch_number,
            "package": item.package,
            "unit": item.unit,
            "quantity_software": item.quantity_software,
            "mrp": item.mrp,
            "unit_price": item.unit_price,
            "selling_price": item.selling_price,
            "profit_margin": item.profit_margin,
            "manufacturing_date": item.manufacturing_date,
            "expiry_date": item.expiry_date,
            "section_id": item.section_id,
            "section_name": item.section.section_name if item.section else None,
            "rack_name": item.section.rack.rack_number if item.section and item.section.rack else None,
            "status": item.status,
            "modified_by_staff": item.modified_by_staff,
            "modified_by_admin": item.modified_by_admin
        })
    
    return {
        "upload": {
            "id": upload.id,
            "filename": upload.filename,
            "uploaded_by": upload.uploaded_by_staff_name,
            "uploaded_at": upload.uploaded_at,
            "status": upload.status,
            "upload_notes": upload.upload_notes,
            "staff_notes": upload.staff_notes,
            "admin_notes": upload.admin_notes,
            "staff_verified": upload.staff_verified,
            "staff_verified_by": upload.staff_verified_by_name,
            "staff_verified_at": upload.staff_verified_at,
            "admin_verified": upload.admin_verified,
            "admin_verified_by": upload.admin_verified_by_name,
            "admin_verified_at": upload.admin_verified_at,
            "rejection_reason": upload.rejection_reason
        },
        "items": result
    }

@router.put("/uploads/{upload_id}/items/{item_id}")
def update_upload_item(
    upload_id: int,
    item_id: int,
    item_data: dict,
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Update item in upload (admin can modify)"""
    upload = db.query(models.ExcelUpload).join(Shop, models.ExcelUpload.shop_id == Shop.id).filter(
        models.ExcelUpload.id == upload_id,
        Shop.organization_id == admin.organization_id
    ).first()
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    item = db.query(models.ExcelUploadItem).filter(
        models.ExcelUploadItem.id == item_id,
        models.ExcelUploadItem.upload_id == upload_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Upload item not found")
    
    for key, value in item_data.items():
        if hasattr(item, key) and key not in ['id', 'upload_id', 'shop_id']:
            if key in ['unit_price', 'selling_price', 'profit_margin', 'quantity_software', 'section_id']:
                if value == '' or value is None:
                    setattr(item, key, None)
                else:
                    try:
                        if key == 'section_id':
                            setattr(item, key, int(value) if value else None)
                        elif key == 'quantity_software':
                            setattr(item, key, int(value) if value else 0)
                        else:
                            setattr(item, key, float(value) if value else None)
                    except (ValueError, TypeError):
                        setattr(item, key, None)
            elif key in ['expiry_date', 'manufacturing_date']:
                setattr(item, key, value if value else None)
            else:
                setattr(item, key, value)

    item.modified_by_admin = True
    db.commit()
    return {"message": "Item updated successfully"}

@router.delete("/uploads/{upload_id}/items/{item_id}")
def delete_upload_item(
    upload_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Delete item from upload (admin)"""
    upload = db.query(models.ExcelUpload).join(Shop, models.ExcelUpload.shop_id == Shop.id).filter(
        models.ExcelUpload.id == upload_id,
        Shop.organization_id == admin.organization_id
    ).first()
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    item = db.query(models.ExcelUploadItem).filter(
        models.ExcelUploadItem.id == item_id,
        models.ExcelUploadItem.upload_id == upload_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Upload item not found")
    
    db.delete(item)
    upload.success_count -= 1
    db.commit()
    return {"message": "Item deleted successfully"}

@router.post("/uploads/{upload_id}/admin-verify")
def admin_verify_upload(
    upload_id: int,
    request_data: dict,
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Admin verification and final approval of upload"""
    notes = request_data.get('notes')
    upload = db.query(models.ExcelUpload).join(Shop, models.ExcelUpload.shop_id == Shop.id).filter(
        models.ExcelUpload.id == upload_id,
        Shop.organization_id == admin.organization_id
    ).first()
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    if upload.status != "pending_admin_verification":
        raise HTTPException(status_code=400, detail="Upload not in correct status for admin verification")
    
    try:
        # Create actual stock items from upload items
        items = db.query(models.ExcelUploadItem).filter(
            models.ExcelUploadItem.upload_id == upload_id,
            models.ExcelUploadItem.status == "pending"
        ).all()
        
        created_count = 0
        for item in items:
            stock_item = models.StockItem(
                shop_id=upload.shop_id,
                product_name=item.product_name,
                batch_number=item.batch_number,
                quantity_software=item.quantity_software,
                composition=item.composition,
                manufacturer=item.manufacturer,
                hsn_code=item.hsn_code,
                package=item.package,
                unit=item.unit,
                expiry_date=item.expiry_date,
                manufacturing_date=item.manufacturing_date,
                mrp=item.mrp,
                unit_price=item.unit_price,
                selling_price=item.selling_price,
                profit_margin=item.profit_margin,
                section_id=item.section_id
            )
            db.add(stock_item)
            db.flush()
            
            item.stock_item_id = stock_item.id
            item.status = "approved"
            created_count += 1
        
        upload.admin_verified = True
        upload.admin_verified_by_id = admin.id
        upload.admin_verified_by_name = admin.full_name
        upload.admin_verified_at = datetime.now()
        upload.admin_notes = notes
        upload.status = "approved"
        
        db.commit()
        return {
            "message": f"Upload approved by admin. {created_count} items added to inventory.",
            "created_count": created_count
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to approve upload: {str(e)}")

@router.post("/uploads/{upload_id}/reject")
def reject_upload(
    upload_id: int,
    request_data: dict,
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Reject upload (admin)"""
    reason = request_data.get('reason')

    upload = db.query(models.ExcelUpload).join(Shop, models.ExcelUpload.shop_id == Shop.id).filter(
        models.ExcelUpload.id == upload_id,
        Shop.organization_id == admin.organization_id
    ).first()
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    upload.status = "rejected"
    upload.rejection_reason = reason
    
    db.commit()
    return {"message": "Upload rejected"}

@router.delete("/uploads/{upload_id}")
def delete_upload(
    upload_id: int,
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Delete upload and associated data (admin only)"""
    upload = db.query(models.ExcelUpload).join(Shop, models.ExcelUpload.shop_id == Shop.id).filter(
        models.ExcelUpload.id == upload_id,
        Shop.organization_id == admin.organization_id
    ).first()
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    # If approved, also delete created stock items
    if upload.status == "approved":
        items = db.query(models.ExcelUploadItem).filter(
            models.ExcelUploadItem.upload_id == upload_id,
            models.ExcelUploadItem.stock_item_id.isnot(None)
        ).all()
        
        for item in items:
            if item.stock_item_id:
                stock_item = db.query(models.StockItem).filter(
                    models.StockItem.id == item.stock_item_id
                ).first()
                if stock_item:
                    db.delete(stock_item)
    
    db.delete(upload)
    db.commit()
    return {"message": "Upload and associated data deleted successfully"}

# ADMIN STOCK ITEMS VIEW

@router.get("/items")
def get_admin_stock_items(
    shop_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    composition: Optional[str] = Query(None),
    manufacturer: Optional[str] = Query(None),
    section_id: Optional[int] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Get paginated stock items for admin (org-scoped, filterable by shop)"""
    import math
    query = (
        db.query(models.StockItem)
        .join(Shop, models.StockItem.shop_id == Shop.id)
        .filter(Shop.organization_id == admin.organization_id)
    )
    if shop_id:
        query = query.filter(models.StockItem.shop_id == shop_id)
    if search:
        s = f"%{search}%"
        query = query.filter(
            (models.StockItem.product_name.ilike(s)) |
            (models.StockItem.batch_number.ilike(s)) |
            (models.StockItem.composition.ilike(s))
        )
    if composition:
        query = query.filter(models.StockItem.composition.ilike(f"%{composition}%"))
    if manufacturer:
        query = query.filter(models.StockItem.manufacturer.ilike(f"%{manufacturer}%"))
    if section_id:
        query = query.filter(models.StockItem.section_id == section_id)

    total = query.count()
    items = query.order_by(models.StockItem.product_name).offset((page - 1) * per_page).limit(per_page).all()

    shops = db.query(Shop).filter(Shop.organization_id == admin.organization_id).all()
    shop_map = {s.id: s.shop_name for s in shops}

    result = []
    for item in items:
        result.append({
            "id": item.id,
            "shop_id": item.shop_id,
            "shop_name": shop_map.get(item.shop_id, "Unknown"),
            "product_name": item.product_name,
            "composition": item.composition,
            "manufacturer": item.manufacturer,
            "hsn_code": item.hsn_code,
            "batch_number": item.batch_number,
            "package": item.package,
            "unit": item.unit,
            "quantity_software": item.quantity_software,
            "quantity_physical": item.quantity_physical,
            "audit_discrepancy": item.audit_discrepancy,
            "mrp": item.mrp,
            "unit_price": item.unit_price,
            "selling_price": item.selling_price,
            "profit_margin": item.profit_margin,
            "manufacturing_date": item.manufacturing_date,
            "expiry_date": item.expiry_date,
            "section_name": item.section.section_name if item.section else None,
            "rack_number": item.section.rack.rack_number if item.section and item.section.rack else None,
            "last_audit_date": item.last_audit_date,
            "created_at": item.created_at
        })

    return {
        "items": result,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": math.ceil(total / per_page) if total > 0 else 1
    }


@router.get("/items/consolidated")
def get_consolidated_stock_items(
    shop_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Get stock items grouped by product_name + composition with per-batch breakdown (org-scoped)"""
    from collections import defaultdict
    query = (
        db.query(models.StockItem)
        .join(Shop, models.StockItem.shop_id == Shop.id)
        .filter(Shop.organization_id == admin.organization_id)
    )
    if shop_id:
        query = query.filter(models.StockItem.shop_id == shop_id)
    if search:
        s = f"%{search}%"
        query = query.filter(
            (models.StockItem.product_name.ilike(s)) |
            (models.StockItem.composition.ilike(s))
        )

    shops = db.query(Shop).filter(Shop.organization_id == admin.organization_id).all()
    shop_map = {s.id: s.shop_name for s in shops}

    items = query.order_by(
        models.StockItem.product_name,
        models.StockItem.composition,
        models.StockItem.batch_number
    ).all()

    groups = defaultdict(lambda: {
        "batches": [], "total_qty_software": 0, "total_qty_physical": 0
    })
    for item in items:
        key = (item.product_name or '', item.composition or '', item.shop_id)
        g = groups[key]
        g["product_name"] = item.product_name
        g["composition"] = item.composition
        g["shop_id"] = item.shop_id
        g["shop_name"] = shop_map.get(item.shop_id, "Unknown")
        g["total_qty_software"] += (item.quantity_software or 0)
        g["total_qty_physical"] += (item.quantity_physical or 0)
        g["batches"].append({
            "batch_number": item.batch_number,
            "expiry_date": item.expiry_date.isoformat() if item.expiry_date else None,
            "qty_software": item.quantity_software or 0,
            "qty_physical": item.quantity_physical or 0,
        })

    result = sorted(groups.values(), key=lambda x: (x.get("product_name") or "").lower())
    for r in result:
        r["batch_count"] = len(r["batches"])
    total = len(result)
    start = (page - 1) * per_page
    return {
        "items": result[start: start + per_page],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": math.ceil(total / per_page) if total > 0 else 1
    }


# ADMIN REPORTS

@router.get("/reports/low-stock")
def get_admin_low_stock(
    threshold: int = Query(10),
    shop_id: Optional[int] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Get low stock items for admin (org-scoped)"""
    import math
    query = (
        db.query(models.StockItem)
        .join(Shop, models.StockItem.shop_id == Shop.id)
        .filter(
            Shop.organization_id == admin.organization_id,
            models.StockItem.quantity_software <= threshold
        )
    )
    if shop_id:
        query = query.filter(models.StockItem.shop_id == shop_id)

    total = query.count()
    items = query.order_by(models.StockItem.quantity_software).offset((page - 1) * per_page).limit(per_page).all()

    shops = db.query(Shop).filter(Shop.organization_id == admin.organization_id).all()
    shop_map = {s.id: s.shop_name for s in shops}

    result = [
        {
            "id": i.id,
            "product_name": i.product_name,
            "batch_number": i.batch_number,
            "quantity_software": i.quantity_software,
            "composition": i.composition,
            "manufacturer": i.manufacturer,
            "expiry_date": i.expiry_date,
            "unit_price": i.unit_price,
            "mrp": i.mrp,
            "shop_name": shop_map.get(i.shop_id, "Unknown")
        }
        for i in items
    ]

    return {
        "items": result,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": math.ceil(total / per_page) if total > 0 else 1
    }


@router.get("/reports/expiring")
def get_admin_expiring(
    days_ahead: int = Query(30),
    shop_id: Optional[int] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Get expiring items for admin (org-scoped)"""
    import math
    cutoff = date.today() + timedelta(days=days_ahead)
    query = (
        db.query(models.StockItem)
        .join(Shop, models.StockItem.shop_id == Shop.id)
        .filter(
            Shop.organization_id == admin.organization_id,
            models.StockItem.expiry_date.isnot(None),
            models.StockItem.expiry_date <= cutoff
        )
    )
    if shop_id:
        query = query.filter(models.StockItem.shop_id == shop_id)

    total = query.count()
    items = query.order_by(models.StockItem.expiry_date).offset((page - 1) * per_page).limit(per_page).all()

    shops = db.query(Shop).filter(Shop.organization_id == admin.organization_id).all()
    shop_map = {s.id: s.shop_name for s in shops}

    result = [
        {
            "id": i.id,
            "product_name": i.product_name,
            "batch_number": i.batch_number,
            "quantity_software": i.quantity_software,
            "expiry_date": i.expiry_date,
            "composition": i.composition,
            "manufacturer": i.manufacturer,
            "mrp": i.mrp,
            "unit_price": i.unit_price,
            "shop_name": shop_map.get(i.shop_id, "Unknown")
        }
        for i in items
    ]

    return {
        "items": result,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": math.ceil(total / per_page) if total > 0 else 1
    }


@router.get("/audit/discrepancies")
def get_admin_audit_discrepancies(
    shop_id: Optional[int] = Query(None),
    threshold: int = Query(0),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Get audit discrepancies for admin (org-scoped)"""
    import math
    query = (
        db.query(models.StockAuditRecord)
        .join(Shop, models.StockAuditRecord.shop_id == Shop.id)
        .filter(
            Shop.organization_id == admin.organization_id,
            models.StockAuditRecord.discrepancy != 0
        )
    )
    if shop_id:
        query = query.filter(models.StockAuditRecord.shop_id == shop_id)

    total = query.count()
    records = query.order_by(models.StockAuditRecord.audit_date.desc()).offset((page - 1) * per_page).limit(per_page).all()

    shops = db.query(Shop).filter(Shop.organization_id == admin.organization_id).all()
    shop_map = {s.id: s.shop_name for s in shops}

    result = []
    for r in records:
        result.append({
            "item": {
                "product_name": r.stock_item.product_name if r.stock_item else "",
                "batch_number": r.stock_item.batch_number if r.stock_item else "",
                "composition": r.stock_item.composition if r.stock_item else ""
            },
            "software_qty": r.software_quantity,
            "physical_qty": r.physical_quantity,
            "difference": r.discrepancy,
            "section_name": r.stock_item.section.section_name if r.stock_item and r.stock_item.section else "",
            "rack_number": r.stock_item.section.rack.rack_number if r.stock_item and r.stock_item.section and r.stock_item.section.rack else "",
            "audited_by_staff_name": r.staff_name,
            "audit_date": r.audit_date,
            "shop_name": shop_map.get(r.shop_id, "Unknown")
        })

    return {
        "discrepancies": result,
        "total_discrepancies": total,
        "page": page,
        "per_page": per_page,
        "pages": math.ceil(total / per_page) if total > 0 else 1
    }


# ADMIN ANALYTICS

@router.get("/analytics/ai-analytics")
def get_admin_ai_analytics(
    shop_id: Optional[int] = Query(None),
    days: int = Query(30),
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Get comprehensive AI analytics for admin tab (same format as staff AI Analytics tab)"""
    import json
    import logging
    from datetime import timedelta

    logger = logging.getLogger(__name__)
    cutoff_date = datetime.now() - timedelta(days=days)

    # Build base queries scoped to org
    item_query = (
        db.query(models.StockItem)
        .join(Shop, models.StockItem.shop_id == Shop.id)
        .filter(Shop.organization_id == admin.organization_id)
    )
    audit_query = (
        db.query(models.StockAuditRecord)
        .join(Shop, models.StockAuditRecord.shop_id == Shop.id)
        .filter(
            Shop.organization_id == admin.organization_id,
            models.StockAuditRecord.audit_date >= cutoff_date
        )
    )
    if shop_id:
        item_query = item_query.filter(models.StockItem.shop_id == shop_id)
        audit_query = audit_query.filter(models.StockAuditRecord.shop_id == shop_id)

    items = item_query.all()
    audit_records = audit_query.all()

    # Build analytics data
    discrepancy_by_date = {}
    for record in audit_records:
        date_key = record.audit_date.date().isoformat()
        if date_key not in discrepancy_by_date:
            discrepancy_by_date[date_key] = {"total": 0, "count": 0}
        discrepancy_by_date[date_key]["total"] += abs(record.discrepancy)
        discrepancy_by_date[date_key]["count"] += 1

    section_discrepancies = {}
    for item in items:
        if item.section and item.audit_discrepancy and item.audit_discrepancy != 0:
            section_name = item.section.section_name
            section_discrepancies[section_name] = section_discrepancies.get(section_name, 0) + abs(item.audit_discrepancy)

    staff_audits = {}
    for record in audit_records:
        staff = record.staff_name or "Unknown"
        if staff not in staff_audits:
            staff_audits[staff] = {"audits": 0, "discrepancies": 0}
        staff_audits[staff]["audits"] += 1
        staff_audits[staff]["discrepancies"] += abs(record.discrepancy)

    total_items = len(items)
    items_with_disc = sum(1 for i in items if (i.audit_discrepancy or 0) != 0)
    total_disc_value = sum(abs(i.audit_discrepancy or 0) for i in items)
    audit_completion = (sum(1 for i in items if i.last_audit_date) / total_items * 100) if total_items > 0 else 0

    analytics_data = {
        "total_audits": len(audit_records),
        "total_items": total_items,
        "items_with_discrepancies": items_with_disc,
        "total_discrepancy_value": total_disc_value,
        "audit_completion_rate": audit_completion,
        "section_discrepancies": section_discrepancies,
        "staff_performance": staff_audits,
    }

    # Build chart data
    trend_dates = sorted(discrepancy_by_date.keys())
    charts = {
        "discrepancy_trend": {
            "labels": trend_dates,
            "datasets": [{
                "label": "Average Discrepancy",
                "data": [discrepancy_by_date[d]["total"] / discrepancy_by_date[d]["count"] for d in trend_dates]
            }]
        },
        "section_discrepancies": {
            "labels": list(section_discrepancies.keys()),
            "data": list(section_discrepancies.values())
        },
        "staff_performance": {
            "labels": list(staff_audits.keys()),
            "datasets": [
                {"label": "Audits Completed", "data": [staff_audits[s]["audits"] for s in staff_audits]},
                {"label": "Discrepancies Found", "data": [staff_audits[s]["discrepancies"] for s in staff_audits]}
            ]
        }
    }

    # Generate AI insights
    ai_insights = None
    try:
        from modules.stock_audit_v2.staff.staff_ai_service import StockAuditAIService
        client = StockAuditAIService._init_gemini()
        if client:
            prompt = f"""
Analyze this pharmacy stock audit data and provide actionable insights:

Summary:
- Total audits conducted: {analytics_data['total_audits']}
- Total items in inventory: {analytics_data['total_items']}
- Items with discrepancies: {analytics_data['items_with_discrepancies']}
- Total discrepancy value: {analytics_data['total_discrepancy_value']} units
- Audit completion rate: {analytics_data['audit_completion_rate']:.1f}%

Section-wise discrepancies:
{json.dumps(section_discrepancies, indent=2)}

Staff performance:
{json.dumps(staff_audits, indent=2)}

Provide:
1. Key findings (3-5 bullet points)
2. Risk areas that need attention
3. Actionable recommendations (3-5 specific actions)
4. Predicted issues if trends continue

Format as JSON with keys: findings, risks, recommendations, predictions
"""
            response = client.models.generate_content(model='gemini-2.0-flash', contents=prompt)
            ai_insights = json.loads(response.text.strip().replace("```json", "").replace("```", ""))
    except Exception as e:
        logger.error(f"Admin AI generation failed: {e}")

    if not ai_insights:
        # Fallback insights
        findings, risks, recommendations = [], [], []
        if analytics_data['total_audits'] == 0:
            findings.append("No audits conducted in the selected period")
            recommendations.append("Start conducting regular stock audits")
        else:
            findings.append(f"Completed {analytics_data['total_audits']} audits covering {total_items} items")

        if items_with_disc > 0:
            rate = (items_with_disc / total_items * 100) if total_items > 0 else 0
            findings.append(f"{rate:.1f}% of items have discrepancies")
            if rate > 20:
                risks.append("High discrepancy rate indicates potential inventory management issues")
                recommendations.append("Implement stricter inventory controls and staff training")

        if audit_completion < 100:
            risks.append(f"Only {audit_completion:.1f}% of items have been audited")
            recommendations.append("Complete audits for all inventory items")

        if section_discrepancies:
            worst = max(section_discrepancies.items(), key=lambda x: x[1])
            risks.append(f"Section '{worst[0]}' has highest discrepancies ({worst[1]} units)")
            recommendations.append(f"Focus audit efforts on {worst[0]} section")

        ai_insights = {
            "findings": findings or ["Insufficient data for analysis"],
            "risks": risks or ["No significant risks identified"],
            "recommendations": recommendations or ["Continue regular audits"],
            "predictions": ["AI predictions unavailable - using statistical analysis"]
        }

    return {
        "summary": {
            "total_audits": analytics_data["total_audits"],
            "total_items": analytics_data["total_items"],
            "items_with_discrepancies": analytics_data["items_with_discrepancies"],
            "total_discrepancy_value": analytics_data["total_discrepancy_value"],
            "audit_completion_rate": round(analytics_data["audit_completion_rate"], 2)
        },
        "charts": charts,
        "ai_insights": ai_insights,
        "generated_at": datetime.now().isoformat()
    }


@router.get("/analytics/dashboard")
def get_admin_dashboard_analytics(
    shop_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Get comprehensive stock audit analytics for admin dashboard"""
    cache_key = f"stock_audit_dashboard:{admin.organization_id}:{shop_id or 'all'}"
    cached = dashboard_cache.get(cache_key)
    if cached is not None:
        return cached
    
    try:
        result = StockAuditAnalytics.get_comprehensive_analytics(
            db=db,
            organization_id=admin.organization_id,
            shop_id=shop_id
        )
        dashboard_cache.set(cache_key, result, ttl=60)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/analytics/ai-insights")
def get_admin_ai_insights(
    shop_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Get AI-generated insights for admin (organization-wide or specific shop)"""
    cache_key = f"stock_audit_ai:{admin.organization_id}:{shop_id or 'all'}"
    cached = dashboard_cache.get(cache_key)
    if cached is not None:
        return cached
    
    try:
        ai_service = StockAuditAIAnalytics()
        result = ai_service.generate_comprehensive_analysis(
            db=db,
            organization_id=admin.organization_id,
            shop_id=shop_id
        )
        dashboard_cache.set(cache_key, result, ttl=3600)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ADMIN EXPORTS

@router.get("/export/stock-items")
def export_stock_items_excel(
    shop_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Export stock items to Excel (admin - all shops or specific shop)"""
    from modules.auth.models import Shop
    
    query = db.query(models.StockItem).join(Shop).filter(
        Shop.organization_id == admin.organization_id
    )
    if shop_id:
        query = query.filter(models.StockItem.shop_id == shop_id)
    
    items = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Items"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ["ID", "Shop", "Product Name", "Composition", "Manufacturer", "HSN Code", "Batch Number", 
               "Package", "Unit", "Rack", "Section", "Software Qty", "Physical Qty", "Discrepancy", 
               "MRP", "Unit Price", "Selling Price", "Profit Margin %", "Total Value", 
               "Mfg Date", "Expiry Date", "Last Audit Date", "Created At"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Get shop names
    shops = db.query(Shop).filter(Shop.organization_id == admin.organization_id).all()
    shop_map = {s.id: s.shop_name for s in shops}
    
    # Data rows
    for row, item in enumerate(items, 2):
        ws.cell(row=row, column=1, value=item.id)
        ws.cell(row=row, column=2, value=shop_map.get(item.shop_id, "Unknown"))
        ws.cell(row=row, column=3, value=item.product_name)
        ws.cell(row=row, column=4, value=item.composition)
        ws.cell(row=row, column=5, value=item.manufacturer)
        ws.cell(row=row, column=6, value=item.hsn_code)
        ws.cell(row=row, column=7, value=item.batch_number)
        ws.cell(row=row, column=8, value=item.package)
        ws.cell(row=row, column=9, value=item.unit)
        ws.cell(row=row, column=10, value=item.section.rack.rack_number if item.section and item.section.rack else "")
        ws.cell(row=row, column=11, value=item.section.section_name if item.section else "")
        ws.cell(row=row, column=12, value=item.quantity_software)
        ws.cell(row=row, column=13, value=item.quantity_physical)
        ws.cell(row=row, column=14, value=item.audit_discrepancy)
        ws.cell(row=row, column=15, value=item.mrp)
        ws.cell(row=row, column=16, value=item.unit_price)
        ws.cell(row=row, column=17, value=item.selling_price)
        ws.cell(row=row, column=18, value=item.profit_margin)
        ws.cell(row=row, column=19, value=(item.quantity_software * item.unit_price) if item.unit_price else None)
        ws.cell(row=row, column=20, value=item.manufacturing_date.strftime("%Y-%m-%d") if item.manufacturing_date else "")
        ws.cell(row=row, column=21, value=item.expiry_date.strftime("%Y-%m-%d") if item.expiry_date else "")
        ws.cell(row=row, column=22, value=item.last_audit_date.strftime("%Y-%m-%d %H:%M") if item.last_audit_date else "")
        ws.cell(row=row, column=23, value=item.created_at.strftime("%Y-%m-%d %H:%M"))
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"stock_items_admin_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export/audit-records")
def export_audit_records_excel(
    days: int = Query(30, description="Number of days to export"),
    shop_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Export audit records to Excel (admin)"""
    from modules.auth.models import Shop
    
    start_date = datetime.now() - timedelta(days=days)
    query = db.query(models.StockAuditRecord).join(Shop).filter(
        Shop.organization_id == admin.organization_id,
        models.StockAuditRecord.audit_date >= start_date
    )
    if shop_id:
        query = query.filter(models.StockAuditRecord.shop_id == shop_id)
    
    records = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Records"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ["ID", "Shop", "Item Name", "Batch Number", "Rack", "Section", "Audit Date", "Staff Name",
               "Software Qty", "Physical Qty", "Discrepancy", "Notes", "Resolved", "Resolution Notes"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Get shop names
    shops = db.query(Shop).filter(Shop.organization_id == admin.organization_id).all()
    shop_map = {s.id: s.shop_name for s in shops}
    
    for row, record in enumerate(records, 2):
        ws.cell(row=row, column=1, value=record.id)
        ws.cell(row=row, column=2, value=shop_map.get(record.shop_id, "Unknown"))
        ws.cell(row=row, column=3, value=record.stock_item.product_name if record.stock_item else "")
        ws.cell(row=row, column=4, value=record.stock_item.batch_number if record.stock_item else "")
        ws.cell(row=row, column=5, value=record.stock_item.section.rack.rack_number if record.stock_item and record.stock_item.section and record.stock_item.section.rack else "")
        ws.cell(row=row, column=6, value=record.stock_item.section.section_name if record.stock_item and record.stock_item.section else "")
        ws.cell(row=row, column=7, value=record.audit_date.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=8, value=record.staff_name)
        ws.cell(row=row, column=9, value=record.software_quantity)
        ws.cell(row=row, column=10, value=record.physical_quantity)
        ws.cell(row=row, column=11, value=record.discrepancy)
        ws.cell(row=row, column=12, value=record.notes)
        ws.cell(row=row, column=13, value="Yes" if record.resolved else "No")
        ws.cell(row=row, column=14, value=record.resolution_notes)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"audit_records_admin_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export/adjustments")
def export_adjustments_excel(
    days: int = Query(30, description="Number of days to export"),
    shop_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Export stock adjustments to Excel (admin)"""
    from modules.auth.models import Shop
    
    start_date = datetime.now() - timedelta(days=days)
    query = db.query(models.StockAdjustment).join(Shop).filter(
        Shop.organization_id == admin.organization_id,
        models.StockAdjustment.adjustment_date >= start_date
    )
    if shop_id:
        query = query.filter(models.StockAdjustment.shop_id == shop_id)
    
    adjustments = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Adjustments"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ["ID", "Shop", "Item Name", "Batch Number", "Rack", "Section", "Adjustment Type", 
               "Quantity Change", "Reason", "Notes", "Staff Name", "Adjustment Date"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Get shop names
    shops = db.query(Shop).filter(Shop.organization_id == admin.organization_id).all()
    shop_map = {s.id: s.shop_name for s in shops}
    
    for row, adj in enumerate(adjustments, 2):
        ws.cell(row=row, column=1, value=adj.id)
        ws.cell(row=row, column=2, value=shop_map.get(adj.shop_id, "Unknown"))
        ws.cell(row=row, column=3, value=adj.stock_item.product_name if adj.stock_item else "")
        ws.cell(row=row, column=4, value=adj.stock_item.batch_number if adj.stock_item else "")
        ws.cell(row=row, column=5, value=adj.stock_item.section.rack.rack_number if adj.stock_item and adj.stock_item.section and adj.stock_item.section.rack else "")
        ws.cell(row=row, column=6, value=adj.stock_item.section.section_name if adj.stock_item and adj.stock_item.section else "")
        ws.cell(row=row, column=7, value=adj.adjustment_type)
        ws.cell(row=row, column=8, value=adj.quantity_change)
        ws.cell(row=row, column=9, value=adj.reason)
        ws.cell(row=row, column=10, value=adj.notes)
        ws.cell(row=row, column=11, value=adj.staff_name)
        ws.cell(row=row, column=12, value=adj.adjustment_date.strftime("%Y-%m-%d %H:%M"))
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"stock_adjustments_admin_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
