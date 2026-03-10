from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database.database import get_db
from datetime import datetime, date, timedelta
from typing import Optional, List
from . import schemas, models, services
from .ai_service import StockAuditAIService
from .admin_analytics import StockAuditAnalytics
from .admin_ai_analytics import StockAuditAIAnalytics
from .dependencies import get_current_user_with_geofence as get_current_user
from modules.auth.dependencies import get_current_admin
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from app.utils.cache import dashboard_cache

router = APIRouter()

# RACK AND SECTION MANAGEMENT

@router.post("/racks", response_model=schemas.StoreRack)
def create_rack(
    rack: schemas.StoreRackCreate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Create a new store rack"""
    staff, shop_id = current_user
    existing = db.query(models.StockRack).filter(
        models.StockRack.rack_number == rack.rack_number,
        models.StockRack.shop_id == shop_id
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="Rack number already exists")
    
    db_rack = models.StockRack(**rack.model_dump(), shop_id=shop_id)
    db.add(db_rack)
    db.commit()
    db.refresh(db_rack)
    return db_rack

@router.get("/racks", response_model=List[schemas.StoreRack])
def get_racks(
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get all store racks"""
    staff, shop_id = current_user
    return db.query(models.StockRack).filter(models.StockRack.shop_id == shop_id).all()

@router.put("/racks/{rack_id}", response_model=schemas.StoreRack)
def update_rack(
    rack_id: int,
    rack: schemas.StoreRackCreate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Update store rack"""
    staff, shop_id = current_user
    db_rack = db.query(models.StockRack).filter(
        models.StockRack.id == rack_id,
        models.StockRack.shop_id == shop_id
    ).first()
    if not db_rack:
        raise HTTPException(status_code=404, detail="Rack not found")
    
    for key, value in rack.model_dump().items():
        setattr(db_rack, key, value)
    
    db.commit()
    db.refresh(db_rack)
    return db_rack

@router.delete("/racks/{rack_id}")
def delete_rack(
    rack_id: int,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Delete store rack"""
    staff, shop_id = current_user
    db_rack = db.query(models.StockRack).filter(
        models.StockRack.id == rack_id,
        models.StockRack.shop_id == shop_id
    ).first()
    if not db_rack:
        raise HTTPException(status_code=404, detail="Rack not found")
    
    db.delete(db_rack)
    db.commit()
    return {"message": "Rack deleted successfully"}

@router.post("/sections", response_model=schemas.StoreSection)
def create_section(
    section: schemas.StoreSectionCreate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Create a new store section"""
    staff, shop_id = current_user
    existing = db.query(models.StockSection).filter(
        models.StockSection.section_code == section.section_code,
        models.StockSection.shop_id == shop_id
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="Section code already exists")
    
    db_section = models.StockSection(**section.model_dump(), shop_id=shop_id)
    db.add(db_section)
    db.commit()
    db.refresh(db_section)
    return db_section

@router.get("/sections", response_model=List[schemas.StoreSection])
def get_sections(
    rack_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get all sections, optionally filtered by rack"""
    staff, shop_id = current_user
    query = db.query(models.StockSection).filter(models.StockSection.shop_id == shop_id)
    if rack_id:
        query = query.filter(models.StockSection.rack_id == rack_id)
    return query.all()

@router.put("/sections/{section_id}", response_model=schemas.StoreSection)
def update_section(
    section_id: int,
    section: schemas.StoreSectionCreate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Update store section"""
    staff, shop_id = current_user
    db_section = db.query(models.StockSection).filter(
        models.StockSection.id == section_id,
        models.StockSection.shop_id == shop_id
    ).first()
    if not db_section:
        raise HTTPException(status_code=404, detail="Section not found")
    
    for key, value in section.model_dump().items():
        setattr(db_section, key, value)
    
    db.commit()
    db.refresh(db_section)
    return db_section

@router.delete("/sections/{section_id}")
def delete_section(
    section_id: int,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Delete store section"""
    staff, shop_id = current_user
    db_section = db.query(models.StockSection).filter(
        models.StockSection.id == section_id,
        models.StockSection.shop_id == shop_id
    ).first()
    if not db_section:
        raise HTTPException(status_code=404, detail="Section not found")
    
    db.delete(db_section)
    db.commit()
    return {"message": "Section deleted successfully"}

# STOCK ITEM MANAGEMENT

@router.post("/items", response_model=schemas.StockItem)
def add_stock_item(
    item: schemas.StockItemCreate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Add new stock item"""
    staff, shop_id = current_user
    db_item = models.StockItem(**item.model_dump(), shop_id=shop_id)
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    return db_item

@router.get("/items", response_model=List[schemas.StockItem])
def get_stock_items(
    section_id: Optional[int] = None,
    rack_id: Optional[int] = None,
    item_name: Optional[str] = None,
    composition: Optional[str] = None,
    batch_number: Optional[str] = None,
    manufacturer: Optional[str] = None,
    expiry_before: Optional[date] = None,
    expiry_after: Optional[date] = None,
    skip: int = 0,
    limit: int = 1000,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get stock items with filters"""
    staff, shop_id = current_user
    query = db.query(models.StockItem).filter(models.StockItem.shop_id == shop_id)
    
    if section_id:
        query = query.filter(models.StockItem.section_id == section_id)
    if rack_id:
        query = query.join(models.StockSection).filter(models.StockSection.rack_id == rack_id)
    if item_name:
        query = query.filter(models.StockItem.product_name.ilike(f"%{item_name}%"))
    if composition:
        query = query.filter(models.StockItem.composition.ilike(f"%{composition}%"))
    if batch_number:
        query = query.filter(models.StockItem.batch_number.ilike(f"%{batch_number}%"))
    if manufacturer:
        query = query.filter(models.StockItem.manufacturer.ilike(f"%{manufacturer}%"))
    if expiry_before:
        query = query.filter(models.StockItem.expiry_date <= expiry_before)
    if expiry_after:
        query = query.filter(models.StockItem.expiry_date >= expiry_after)
    
    items = query.offset(skip).limit(limit).all()
    
    # Sort alphabetically by product_name
    items = sorted(items, key=lambda x: x.product_name.lower())
    
    # Add section and rack names
    result = []
    for item in items:
        item_dict = {
            **item.__dict__,
            "section_name": item.section.section_name if item.section else None,
            "rack_name": item.section.rack.rack_number if item.section and item.section.rack else None,
            "total_value": (item.quantity_software * item.unit_price) if item.unit_price else None
        }
        result.append(item_dict)
    
    return result

@router.get("/items/{item_id}", response_model=schemas.StockItem)
def get_stock_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get specific stock item"""
    staff, shop_id = current_user
    item = db.query(models.StockItem).filter(
        models.StockItem.id == item_id,
        models.StockItem.shop_id == shop_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Stock item not found")
    
    # Add section and rack names
    item_dict = {
        **item.__dict__,
        "section_name": item.section.section_name if item.section else None,
        "rack_name": item.section.rack.rack_number if item.section and item.section.rack else None,
        "total_value": (item.quantity_software * item.unit_price) if item.unit_price else None
    }
    return item_dict

@router.put("/items/{item_id}", response_model=schemas.StockItem)
def update_stock_item(
    item_id: int,
    item: schemas.StockItemCreate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Update stock item"""
    staff, shop_id = current_user
    db_item = db.query(models.StockItem).filter(
        models.StockItem.id == item_id,
        models.StockItem.shop_id == shop_id
    ).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Stock item not found")
    
    for key, value in item.model_dump().items():
        setattr(db_item, key, value)
    
    db.commit()
    db.refresh(db_item)
    return db_item

@router.delete("/items/{item_id}")
def delete_stock_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Delete stock item"""
    staff, shop_id = current_user
    db_item = db.query(models.StockItem).filter(
        models.StockItem.id == item_id,
        models.StockItem.shop_id == shop_id
    ).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Stock item not found")
    
    db.delete(db_item)
    db.commit()
    return {"message": "Stock item deleted successfully"}

@router.post("/items/bulk-delete")
def bulk_delete_items(
    item_ids: List[int],
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Bulk delete stock items"""
    staff, shop_id = current_user
    try:
        deleted_count = db.query(models.StockItem).filter(
            models.StockItem.id.in_(item_ids),
            models.StockItem.shop_id == shop_id
        ).delete(synchronize_session=False)
        db.commit()
        return {"message": f"{deleted_count} items deleted successfully", "count": deleted_count}
    except Exception as e:
        db.rollback()
        if "foreign key constraint" in str(e).lower():
            raise HTTPException(
                status_code=400, 
                detail="Cannot delete items that have been used in bills or transactions. Please remove them from bills first."
            )
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/items/upload-excel")
async def upload_stock_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Upload stock items from Excel file - matches export format"""
    staff, shop_id = current_user
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed")
    
    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Match export format: ID, Product Name, Composition, Manufacturer, HSN Code, Batch Number, 
        # Package, Unit, Rack, Section, Software Qty, Physical Qty, Discrepancy, 
        # MRP, Unit Price, Selling Price, Profit Margin %, Total Value, Mfg Date, Expiry Date, Last Audit Date, Created At
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[1]:  # Skip empty rows (check product name at index 1)
                continue
                
            try:
                # Skip ID (row[0])
                product_name = str(row[1]).strip() if row[1] else None
                composition = str(row[2]).strip() if len(row) > 2 and row[2] else None
                manufacturer = str(row[3]).strip() if len(row) > 3 and row[3] else None
                hsn_code = str(row[4]).strip() if len(row) > 4 and row[4] else None
                batch_number = str(row[5]).strip() if len(row) > 5 and row[5] else None
                package = str(row[6]).strip() if len(row) > 6 and row[6] else None
                unit = str(row[7]).strip() if len(row) > 7 and row[7] else None
                # Skip Rack (row[8]) and Section (row[9]) - will be assigned later
                quantity = int(row[10]) if len(row) > 10 and row[10] else 0
                # Skip Physical Qty (row[11]) and Discrepancy (row[12])
                mrp = str(row[13]).strip() if len(row) > 13 and row[13] else None
                unit_price = float(row[14]) if len(row) > 14 and row[14] else None
                selling_price = float(row[15]) if len(row) > 15 and row[15] else None
                
                # Handle profit margin - cap at 999.99
                profit_margin = None
                if len(row) > 16 and row[16]:
                    try:
                        pm = float(row[16])
                        profit_margin = min(pm, 999.99)  # Cap at database limit
                    except:
                        pass
                
                # Skip Total Value (row[17])
                
                # Handle dates
                manufacturing_date = None
                if len(row) > 18 and row[18]:
                    try:
                        if isinstance(row[18], datetime):
                            manufacturing_date = row[18].date()
                        elif row[18] and str(row[18]).strip():
                            manufacturing_date = datetime.strptime(str(row[18]), "%Y-%m-%d").date()
                    except:
                        pass
                
                expiry_date = None
                if len(row) > 19 and row[19]:
                    try:
                        if isinstance(row[19], datetime):
                            expiry_date = row[19].date()
                        elif row[19] and str(row[19]).strip():
                            expiry_date = datetime.strptime(str(row[19]), "%Y-%m-%d").date()
                    except:
                        pass
                
                if not product_name or not batch_number:
                    errors.append(f"Row {row_idx}: Product name and batch number are required")
                    error_count += 1
                    continue
                
                # Create stock item (section will be unassigned)
                db_item = models.StockItem(
                    product_name=product_name,
                    batch_number=batch_number,
                    quantity_software=quantity,
                    composition=composition,
                    manufacturer=manufacturer,
                    hsn_code=hsn_code,
                    package=package,
                    unit=unit,
                    expiry_date=expiry_date,
                    manufacturing_date=manufacturing_date,
                    mrp=mrp,
                    unit_price=unit_price,
                    selling_price=selling_price,
                    profit_margin=profit_margin,
                    section_id=None,  # Will be unassigned
                    shop_id=shop_id
                )
                db.add(db_item)
                success_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                error_count += 1
        
        db.commit()
        
        return {
            "message": f"Upload completed: {success_count} items added, {error_count} errors",
            "success_count": success_count,
            "error_count": error_count,
            "errors": errors[:10]  # Return first 10 errors
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to process Excel file: {str(e)}")

@router.get("/items/unassigned/list", response_model=List[schemas.StockItem])
def get_unassigned_items(
    item_name: Optional[str] = None,
    composition: Optional[str] = None,
    manufacturer: Optional[str] = None,
    batch_number: Optional[str] = None,
    skip: int = 0,
    limit: int = 1000,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get stock items without assigned rack/section with search and filters"""
    staff, shop_id = current_user
    query = db.query(models.StockItem).filter(
        models.StockItem.shop_id == shop_id,
        models.StockItem.section_id.is_(None)
    )
    
    if item_name:
        query = query.filter(models.StockItem.product_name.ilike(f"%{item_name}%"))
    if composition:
        query = query.filter(models.StockItem.composition.ilike(f"%{composition}%"))
    if manufacturer:
        query = query.filter(models.StockItem.manufacturer.ilike(f"%{manufacturer}%"))
    if batch_number:
        query = query.filter(models.StockItem.batch_number.ilike(f"%{batch_number}%"))
    
    items = query.offset(skip).limit(limit).all()
    
    # Sort alphabetically by product_name
    items = sorted(items, key=lambda x: x.product_name.lower())
    
    result = []
    for item in items:
        item_dict = {
            **item.__dict__,
            "section_name": None,
            "rack_name": None,
            "total_value": (item.quantity_software * item.unit_price) if item.unit_price else None
        }
        result.append(item_dict)
    
    return result

@router.patch("/items/{item_id}/assign-section")
def assign_section_to_item(
    item_id: int,
    section_id: int,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Assign rack/section to a stock item"""
    staff, shop_id = current_user
    
    # Verify section exists
    section = db.query(models.StockSection).filter(
        models.StockSection.id == section_id,
        models.StockSection.shop_id == shop_id
    ).first()
    if not section:
        raise HTTPException(status_code=404, detail="Section not found")
    
    # Update item
    db_item = db.query(models.StockItem).filter(
        models.StockItem.id == item_id,
        models.StockItem.shop_id == shop_id
    ).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Stock item not found")
    
    db_item.section_id = section_id
    db_item.updated_at = datetime.now()
    db.commit()
    
    return {
        "message": "Section assigned successfully",
        "item_id": item_id,
        "section_id": section_id,
        "section_name": section.section_name,
        "rack_number": section.rack.rack_number
    }

# PURCHASE MANAGEMENT

@router.post("/purchases", response_model=schemas.Purchase)
def add_purchase(
    purchase_data: dict,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Add purchase and update stock levels"""
    staff, shop_id = current_user
    try:
        purchase = services.StockCalculationService.add_purchase(
            db, 
            purchase_data['purchase'], 
            purchase_data['items'],
            shop_id,
            staff.id,
            staff.name
        )
        return purchase
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/purchases", response_model=List[schemas.Purchase])
def get_purchases(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    supplier_name: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get purchases with filters"""
    staff, shop_id = current_user
    query = db.query(models.Purchase).filter(models.Purchase.shop_id == shop_id)
    
    if start_date:
        query = query.filter(models.Purchase.purchase_date >= start_date)
    if end_date:
        query = query.filter(models.Purchase.purchase_date <= end_date)
    if supplier_name:
        query = query.filter(models.Purchase.supplier_name.ilike(f"%{supplier_name}%"))
    
    return query.order_by(models.Purchase.purchase_date.desc()).offset(skip).limit(limit).all()

@router.put("/purchases/{purchase_id}", response_model=schemas.Purchase)
def update_purchase(
    purchase_id: int,
    purchase_data: dict,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Update purchase record"""
    staff, shop_id = current_user
    db_purchase = db.query(models.Purchase).filter(
        models.Purchase.id == purchase_id,
        models.Purchase.shop_id == shop_id
    ).first()
    if not db_purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    for key, value in purchase_data.get('purchase', {}).items():
        if hasattr(db_purchase, key):
            setattr(db_purchase, key, value)
    
    db.commit()
    db.refresh(db_purchase)
    return db_purchase

@router.delete("/purchases/{purchase_id}")
def delete_purchase(
    purchase_id: int,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Delete purchase record"""
    staff, shop_id = current_user
    db_purchase = db.query(models.Purchase).filter(
        models.Purchase.id == purchase_id,
        models.Purchase.shop_id == shop_id
    ).first()
    if not db_purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    db.delete(db_purchase)
    db.commit()
    return {"message": "Purchase deleted successfully"}

# SALES MANAGEMENT

@router.post("/sales", response_model=schemas.Sale)
def add_sale(
    sale_data: dict,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Add sale and update stock levels"""
    staff, shop_id = current_user
    try:
        sale = services.StockCalculationService.add_sale(
            db, 
            sale_data['sale'], 
            sale_data['items'],
            shop_id,
            staff.id,
            staff.name
        )
        return sale
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/sales", response_model=List[schemas.Sale])
def get_sales(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    customer_phone: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get sales with filters"""
    staff, shop_id = current_user
    query = db.query(models.Sale).filter(models.Sale.shop_id == shop_id)
    
    if start_date:
        query = query.filter(models.Sale.sale_date >= start_date)
    if end_date:
        query = query.filter(models.Sale.sale_date <= end_date)
    if customer_phone:
        query = query.filter(models.Sale.customer_phone == customer_phone)
    
    return query.order_by(models.Sale.sale_date.desc()).offset(skip).limit(limit).all()

@router.put("/sales/{sale_id}", response_model=schemas.Sale)
def update_sale(
    sale_id: int,
    sale_data: dict,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Update sale record"""
    staff, shop_id = current_user
    db_sale = db.query(models.Sale).filter(
        models.Sale.id == sale_id,
        models.Sale.shop_id == shop_id
    ).first()
    if not db_sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    
    for key, value in sale_data.get('sale', {}).items():
        if hasattr(db_sale, key):
            setattr(db_sale, key, value)
    
    db.commit()
    db.refresh(db_sale)
    return db_sale

@router.delete("/sales/{sale_id}")
def delete_sale(
    sale_id: int,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Delete sale record"""
    staff, shop_id = current_user
    db_sale = db.query(models.Sale).filter(
        models.Sale.id == sale_id,
        models.Sale.shop_id == shop_id
    ).first()
    if not db_sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    
    db.delete(db_sale)
    db.commit()
    return {"message": "Sale deleted successfully"}

# AUDIT FUNCTIONALITY

@router.get("/audit/random-section", response_model=schemas.RandomAuditSection)
def get_random_audit_section(
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get random section for audit"""
    staff, shop_id = current_user
    result = services.StockAuditService.get_random_section_for_audit(db, shop_id)
    if not result:
        raise HTTPException(status_code=404, detail="No sections available for audit")
    return result

@router.post("/audit/sessions", response_model=schemas.StockAuditSession)
def start_audit_session(
    session: schemas.StockAuditSessionCreate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Start new audit session"""
    staff, shop_id = current_user
    return services.StockAuditService.start_audit_session(db, staff.id, staff.name, shop_id)

@router.put("/items/{item_id}/audit", response_model=schemas.StockAuditRecord)
def audit_stock_item(
    item_id: int,
    physical_quantity: int,
    notes: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Record audit result for stock item"""
    staff, shop_id = current_user
    try:
        audit_record = services.StockAuditService.record_audit(
            db, item_id, physical_quantity, staff.id, staff.name, notes, shop_id
        )
        return audit_record
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/audit/discrepancies")
def get_stock_discrepancies(
    threshold: int = Query(0, description="Minimum discrepancy threshold"),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get all stock discrepancies above threshold"""
    staff, shop_id = current_user
    discrepancies = services.StockAuditService.get_discrepancies(db, threshold, shop_id)
    return {
        "total_discrepancies": len(discrepancies),
        "threshold": threshold,
        "discrepancies": discrepancies
    }

@router.get("/audit/summary")
def get_audit_summary(
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get overall audit summary"""
    staff, shop_id = current_user
    return services.StockAuditService.get_audit_summary(db, shop_id)

# STOCK CALCULATIONS

@router.post("/calculate-stock")
def update_software_stock(
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Recalculate software stock for all items based on purchases/sales"""
    staff, shop_id = current_user
    result = services.StockCalculationService.update_all_software_stock(db, shop_id)
    return {
        "message": "Stock calculations updated",
        "details": result
    }

@router.get("/items/{item_id}/stock-calculation")
def get_item_stock_calculation(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get detailed stock calculation for specific item"""
    staff, shop_id = current_user
    item = db.query(models.StockItem).filter(
        models.StockItem.id == item_id,
        models.StockItem.shop_id == shop_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Stock item not found")
    
    calculated_stock = services.StockCalculationService.calculate_software_stock(db, item_id)
    
    purchases = db.query(models.PurchaseItem).filter(
        models.PurchaseItem.stock_item_id == item_id
    ).all()
    
    sales = db.query(models.SaleItem).filter(
        models.SaleItem.stock_item_id == item_id
    ).all()
    
    total_purchased = sum(p.quantity for p in purchases)
    total_sold = sum(s.quantity for s in sales)
    
    return {
        "item": item,
        "calculated_stock": calculated_stock,
        "current_software_stock": item.quantity_software,
        "total_purchased": total_purchased,
        "total_sold": total_sold,
        "purchase_transactions": len(purchases),
        "sale_transactions": len(sales)
    }

# REPORTS

@router.get("/reports/low-stock")
def get_low_stock_report(
    threshold: int = Query(10, description="Low stock threshold"),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get low stock items report"""
    staff, shop_id = current_user
    items = services.StockReportService.get_low_stock_items(db, threshold, shop_id)
    return {
        "threshold": threshold,
        "total_low_stock_items": len(items),
        "items": items
    }

@router.get("/reports/expiring")
def get_expiring_items_report(
    days_ahead: int = Query(30, description="Days ahead to check for expiry"),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get expiring items report"""
    staff, shop_id = current_user
    items = services.StockReportService.get_expiring_items(db, days_ahead, shop_id)
    return {
        "days_ahead": days_ahead,
        "total_expiring_items": len(items),
        "items": items
    }

@router.get("/reports/stock-movement")
def get_stock_movement_report(
    start_date: date,
    end_date: date,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get stock movement report for date range"""
    staff, shop_id = current_user
    return services.StockReportService.get_stock_movement_report(db, start_date, end_date, shop_id)

@router.post("/adjustments", response_model=schemas.StockAdjustment)
def create_stock_adjustment(
    adjustment: schemas.StockAdjustmentCreate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Create stock adjustment for corrections/damages/returns"""
    staff, shop_id = current_user
    item = db.query(models.StockItem).filter(
        models.StockItem.id == adjustment.stock_item_id,
        models.StockItem.shop_id == shop_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Stock item not found")
    
    db_adjustment = models.StockAdjustment(
        **adjustment.model_dump(),
        shop_id=shop_id,
        staff_id=staff.id,
        staff_name=staff.name
    )
    db.add(db_adjustment)
    
    item.quantity_software += adjustment.quantity_change
    if item.quantity_software < 0:
        raise HTTPException(status_code=400, detail="Adjustment would result in negative stock")
    
    db.commit()
    db.refresh(db_adjustment)
    return db_adjustment

@router.get("/adjustments", response_model=List[schemas.StockAdjustment])
def get_stock_adjustments(
    stock_item_id: Optional[int] = None,
    adjustment_type: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get stock adjustments with filters"""
    staff, shop_id = current_user
    query = db.query(models.StockAdjustment).filter(models.StockAdjustment.shop_id == shop_id)
    
    if stock_item_id:
        query = query.filter(models.StockAdjustment.stock_item_id == stock_item_id)
    if adjustment_type:
        query = query.filter(models.StockAdjustment.adjustment_type == adjustment_type)
    
    return query.order_by(models.StockAdjustment.adjustment_date.desc()).offset(skip).limit(limit).all()

# AI ANALYTICS

@router.get("/admin/analytics/dashboard")
def get_admin_dashboard_analytics(
    shop_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Get comprehensive stock audit analytics for admin dashboard"""
    cache_key = f"stock_audit_dashboard:{admin.organization_id}:{shop_id or 'all'}"
    cached = dashboard_cache.get(cache_key)
    if cached is not None:
        return cached
    
    try:
        result = StockAuditAnalytics.get_comprehensive_analytics(
            db=db,
            organization_id=admin.organization_id,
            shop_id=shop_id
        )
        dashboard_cache.set(cache_key, result, ttl=60)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/admin/analytics/ai-insights")
def get_admin_ai_insights(
    shop_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    admin = Depends(get_current_admin)
):
    """Get AI-generated insights for admin (organization-wide or specific shop)"""
    cache_key = f"stock_audit_ai:{admin.organization_id}:{shop_id or 'all'}"
    cached = dashboard_cache.get(cache_key)
    if cached is not None:
        return cached
    
    try:
        ai_service = StockAuditAIAnalytics()
        result = ai_service.generate_comprehensive_analysis(
            db=db,
            organization_id=admin.organization_id,
            shop_id=shop_id
        )
        dashboard_cache.set(cache_key, result, ttl=3600)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/ai-analytics/comprehensive")
def get_comprehensive_ai_analysis(
    days: int = Query(30, description="Number of days to analyze"),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get comprehensive AI analysis with charts and insights"""
    staff, shop_id = current_user
    try:
        return StockAuditAIService.get_comprehensive_analysis(db, shop_id, days)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/ai-analytics/charts")
def get_ai_chart_data(
    days: int = Query(30, description="Number of days to analyze"),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get chart data for frontend visualization"""
    staff, shop_id = current_user
    return StockAuditAIService.get_chart_data(db, shop_id, days)

@router.get("/ai-analytics/insights")
def get_ai_insights(
    days: int = Query(30, description="Number of days to analyze"),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get AI-generated insights and recommendations"""
    staff, shop_id = current_user
    try:
        return StockAuditAIService.get_ai_insights(db, shop_id, days)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# EXCEL EXPORT

@router.get("/export/stock-items")
def export_stock_items_excel(
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Export all stock items to Excel"""
    staff, shop_id = current_user
    items = db.query(models.StockItem).filter(models.StockItem.shop_id == shop_id).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Items"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ["ID", "Product Name", "Composition", "Manufacturer", "HSN Code", "Batch Number", 
               "Package", "Unit", "Rack", "Section", "Software Qty", "Physical Qty", "Discrepancy", 
               "MRP", "Unit Price", "Selling Price", "Profit Margin %", "Total Value", 
               "Mfg Date", "Expiry Date", "Last Audit Date", "Created At"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for row, item in enumerate(items, 2):
        ws.cell(row=row, column=1, value=item.id)
        ws.cell(row=row, column=2, value=item.product_name)
        ws.cell(row=row, column=3, value=item.composition)
        ws.cell(row=row, column=4, value=item.manufacturer)
        ws.cell(row=row, column=5, value=item.hsn_code)
        ws.cell(row=row, column=6, value=item.batch_number)
        ws.cell(row=row, column=7, value=item.package)
        ws.cell(row=row, column=8, value=item.unit)
        ws.cell(row=row, column=9, value=item.section.rack.rack_number if item.section and item.section.rack else "")
        ws.cell(row=row, column=10, value=item.section.section_name if item.section else "")
        ws.cell(row=row, column=11, value=item.quantity_software)
        ws.cell(row=row, column=12, value=item.quantity_physical)
        ws.cell(row=row, column=13, value=item.audit_discrepancy)
        ws.cell(row=row, column=14, value=item.mrp)
        ws.cell(row=row, column=15, value=item.unit_price)
        ws.cell(row=row, column=16, value=item.selling_price)
        ws.cell(row=row, column=17, value=item.profit_margin)
        ws.cell(row=row, column=18, value=(item.quantity_software * item.unit_price) if item.unit_price else None)
        ws.cell(row=row, column=19, value=item.manufacturing_date.strftime("%Y-%m-%d") if item.manufacturing_date else "")
        ws.cell(row=row, column=20, value=item.expiry_date.strftime("%Y-%m-%d") if item.expiry_date else "")
        ws.cell(row=row, column=21, value=item.last_audit_date.strftime("%Y-%m-%d %H:%M") if item.last_audit_date else "")
        ws.cell(row=row, column=22, value=item.created_at.strftime("%Y-%m-%d %H:%M"))
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"stock_items_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export/audit-records")
def export_audit_records_excel(
    days: int = Query(30, description="Number of days to export"),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Export audit records to Excel"""
    staff, shop_id = current_user
    start_date = datetime.now() - timedelta(days=days)
    records = db.query(models.StockAuditRecord).filter(
        models.StockAuditRecord.shop_id == shop_id,
        models.StockAuditRecord.audit_date >= start_date
    ).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Records"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ["ID", "Item Name", "Batch Number", "Rack", "Section", "Audit Date", "Staff Name",
               "Software Qty", "Physical Qty", "Discrepancy", "Notes", "Resolved", "Resolution Notes"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row, record in enumerate(records, 2):
        ws.cell(row=row, column=1, value=record.id)
        ws.cell(row=row, column=2, value=record.stock_item.item_name if record.stock_item else "")
        ws.cell(row=row, column=3, value=record.stock_item.batch_number if record.stock_item else "")
        ws.cell(row=row, column=4, value=record.stock_item.section.rack.rack_number if record.stock_item and record.stock_item.section and record.stock_item.section.rack else "")
        ws.cell(row=row, column=5, value=record.stock_item.section.section_name if record.stock_item and record.stock_item.section else "")
        ws.cell(row=row, column=6, value=record.audit_date.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=7, value=record.staff_name)
        ws.cell(row=row, column=8, value=record.software_quantity)
        ws.cell(row=row, column=9, value=record.physical_quantity)
        ws.cell(row=row, column=10, value=record.discrepancy)
        ws.cell(row=row, column=11, value=record.notes)
        ws.cell(row=row, column=12, value="Yes" if record.resolved else "No")
        ws.cell(row=row, column=13, value=record.resolution_notes)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"audit_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export/adjustments")
def export_adjustments_excel(
    days: int = Query(30, description="Number of days to export"),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Export stock adjustments to Excel"""
    staff, shop_id = current_user
    start_date = datetime.now() - timedelta(days=days)
    adjustments = db.query(models.StockAdjustment).filter(
        models.StockAdjustment.shop_id == shop_id,
        models.StockAdjustment.adjustment_date >= start_date
    ).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Adjustments"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ["ID", "Item Name", "Batch Number", "Rack", "Section", "Adjustment Type", 
               "Quantity Change", "Reason", "Notes", "Staff Name", "Adjustment Date"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row, adj in enumerate(adjustments, 2):
        ws.cell(row=row, column=1, value=adj.id)
        ws.cell(row=row, column=2, value=adj.stock_item.item_name if adj.stock_item else "")
        ws.cell(row=row, column=3, value=adj.stock_item.batch_number if adj.stock_item else "")
        ws.cell(row=row, column=4, value=adj.stock_item.section.rack.rack_number if adj.stock_item and adj.stock_item.section and adj.stock_item.section.rack else "")
        ws.cell(row=row, column=5, value=adj.stock_item.section.section_name if adj.stock_item and adj.stock_item.section else "")
        ws.cell(row=row, column=6, value=adj.adjustment_type)
        ws.cell(row=row, column=7, value=adj.quantity_change)
        ws.cell(row=row, column=8, value=adj.reason)
        ws.cell(row=row, column=9, value=adj.notes)
        ws.cell(row=row, column=10, value=adj.staff_name)
        ws.cell(row=row, column=11, value=adj.adjustment_date.strftime("%Y-%m-%d %H:%M"))
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"stock_adjustments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
