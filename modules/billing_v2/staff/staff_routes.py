from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database.database import get_db
from datetime import datetime, date, timedelta
from typing import Optional, List
import io
import os
import json
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic import BaseModel
from app.utils.cache import dashboard_cache
from modules.auth.models import Shop
from modules.billing_v2 import schemas, models, services
from modules.billing_v2 import daily_records_schemas
from modules.billing_v2.daily_records_service import DailyRecordsService
from modules.billing_v2.daily_records_models import DailyRecord, DailyExpense
from modules.billing_v2.staff.staff_dependencies import get_current_user_with_geofence as get_current_user

router = APIRouter()

# ─── BILLING USER GUIDE ───────────────────────────────────────────────────────

@router.get("/user-guide")
def get_user_guide():
    """Get billing system user guide"""
    try:
        guide_path = os.path.join(os.path.dirname(__file__), "..", "BILLING_USER_GUIDE.md")
        with open(guide_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return {"content": content}
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="User guide not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading user guide: {str(e)}")

# ─── MEDICINE SEARCH ──────────────────────────────────────────────────────────

@router.get("/search-medicines", response_model=List[schemas.MedicineSearchResult])
def search_medicines(
    q: str = Query(..., min_length=2, description="Search term"),
    limit: int = Query(20, le=100),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Search medicines by name, generic name, brand, or batch number with location info"""
    staff, shop_id = current_user
    results = services.BillingService.search_medicines(db, shop_id, q, limit)
    return results

# ─── BILL MANAGEMENT ──────────────────────────────────────────────────────────

@router.post("/bills", response_model=schemas.BillResponse)
def create_bill(
    bill_data: schemas.BillCreate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Create new bill and update stock automatically"""
    staff, shop_id = current_user

    try:
        bill = services.BillingService.create_bill(
            db,
            shop_id,
            staff.id,
            staff.name,
            bill_data.model_dump(exclude={'items'}),
            [item.model_dump() for item in bill_data.items]
        )
        return bill
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/bills", response_model=List[schemas.BillResponse])
def get_bills(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    customer_phone: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get bills with filters"""
    staff, shop_id = current_user
    query = db.query(models.Bill).filter(models.Bill.shop_id == shop_id)

    if start_date:
        query = query.filter(models.Bill.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.filter(models.Bill.created_at <= datetime.combine(end_date, datetime.max.time()))
    if customer_phone:
        query = query.filter(models.Bill.customer_phone == customer_phone)

    return query.order_by(models.Bill.created_at.desc()).offset(skip).limit(limit).all()

@router.get("/bills/{bill_id}", response_model=schemas.BillResponse)
def get_bill(
    bill_id: int,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get specific bill details"""
    staff, shop_id = current_user
    bill = db.query(models.Bill).filter(
        models.Bill.id == bill_id,
        models.Bill.shop_id == shop_id
    ).first()

    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")

    return bill

@router.get("/bills/number/{bill_number}", response_model=schemas.BillResponse)
def get_bill_by_number(
    bill_number: str,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get bill by bill number"""
    staff, shop_id = current_user
    bill = db.query(models.Bill).filter(
        models.Bill.bill_number == bill_number,
        models.Bill.shop_id == shop_id
    ).first()

    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")

    return bill

@router.delete("/bills/{bill_id}")
def delete_bill(
    bill_id: int,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Delete bill (managers only - restores stock)"""
    staff, shop_id = current_user

    if not staff.can_manage_staff:  # Only managers can delete bills
        raise HTTPException(status_code=403, detail="Permission denied")

    bill = db.query(models.Bill).filter(
        models.Bill.id == bill_id,
        models.Bill.shop_id == shop_id
    ).first()

    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")

    # Restore stock quantities and recalculate discrepancy
    from modules.stock_audit_v2.models import StockItem
    for item in bill.items:
        stock_item = db.query(StockItem).filter(
            StockItem.id == item.stock_item_id,
            StockItem.shop_id == shop_id
        ).first()
        if stock_item:
            # Use strips_deducted if available (new bills); fall back to quantity for legacy bills
            restore_qty = item.strips_deducted if item.strips_deducted is not None else item.quantity
            stock_item.quantity_software += restore_qty
            if stock_item.quantity_physical is not None:
                stock_item.audit_discrepancy = stock_item.quantity_software - stock_item.quantity_physical
            stock_item.updated_at = datetime.now()

    db.delete(bill)
    db.commit()
    return {"message": "Bill deleted and stock restored"}

# ─── PAY LATER ────────────────────────────────────────────────────────────────

@router.get("/pay-later/customers")
def get_pay_later_customers(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get all customers with outstanding Pay Later balances"""
    staff, shop_id = current_user
    return services.BillingService.get_pay_later_customers(db, shop_id, from_date, to_date)

@router.get("/pay-later/bills/{customer_phone}")
def get_pay_later_bills(
    customer_phone: str,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get all outstanding Pay Later bills for a specific customer"""
    staff, shop_id = current_user
    bills = db.query(models.Bill).filter(
        models.Bill.shop_id == shop_id,
        models.Bill.customer_phone == customer_phone,
        models.Bill.payment_status.in_(['pay_later', 'partial'])
    ).order_by(models.Bill.created_at.asc()).all()
    return bills

@router.post("/pay-later/record-payment")
def record_pay_later_payment(
    payment: schemas.RecordPaymentRequest,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Record a payment for a customer's outstanding Pay Later bills (FIFO order)"""
    staff, shop_id = current_user
    try:
        result = services.BillingService.record_payment(db, shop_id, payment.model_dump())
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

# ─── REPORTS ──────────────────────────────────────────────────────────────────

@router.get("/summary", response_model=schemas.BillSummary)
def get_billing_summary(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get billing summary for date range"""
    staff, shop_id = current_user
    return services.BillingService.get_bill_summary(db, shop_id, start_date, end_date)

@router.get("/top-selling")
def get_top_selling_items(
    limit: int = Query(10, le=50),
    days: int = Query(30, le=365),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get top selling items"""
    staff, shop_id = current_user
    return services.BillingService.get_top_selling_items(db, shop_id, limit, days)

@router.get("/customer-history/{customer_phone}")
def get_customer_history(
    customer_phone: str,
    limit: int = Query(50, le=100),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get customer purchase history"""
    staff, shop_id = current_user
    bills = db.query(models.Bill).filter(
        models.Bill.shop_id == shop_id,
        models.Bill.customer_phone == customer_phone
    ).order_by(models.Bill.created_at.desc()).limit(limit).all()

    total_spent = sum(b.total_amount for b in bills)

    return {
        "customer_phone": customer_phone,
        "total_bills": len(bills),
        "total_spent": float(total_spent),
        "bills": bills
    }

@router.get("/daily-sales")
def get_daily_sales(
    days: int = Query(7, le=90),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get daily sales for last N days"""
    staff, shop_id = current_user
    cutoff_date = datetime.now() - timedelta(days=days)

    results = db.query(
        func.date(models.Bill.created_at).label('date'),
        func.count(models.Bill.id).label('bill_count'),
        func.sum(models.Bill.total_amount).label('total_sales')
    ).filter(
        models.Bill.shop_id == shop_id,
        models.Bill.created_at >= cutoff_date
    ).group_by(
        func.date(models.Bill.created_at)
    ).order_by(
        func.date(models.Bill.created_at).desc()
    ).all()

    return [
        {
            "date": str(r.date),
            "bill_count": r.bill_count,
            "total_sales": float(r.total_sales)
        }
        for r in results
    ]

# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────

@router.get("/export/bills")
def export_bills_excel(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Export bills to Excel"""
    staff, shop_id = current_user
    query = db.query(models.Bill).filter(models.Bill.shop_id == shop_id)

    if start_date:
        query = query.filter(models.Bill.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.filter(models.Bill.created_at <= datetime.combine(end_date, datetime.max.time()))

    bills = query.order_by(models.Bill.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bills"

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["Bill Number", "Date", "Customer Name", "Customer Phone", "Doctor Name", "Payment Method",
               "Subtotal", "Discount", "Tax", "Total Amount", "Amount Paid", "Change", "Staff Name"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row, bill in enumerate(bills, 2):
        ws.cell(row=row, column=1, value=bill.bill_number)
        ws.cell(row=row, column=2, value=bill.created_at.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=3, value=bill.customer_name or "")
        ws.cell(row=row, column=4, value=bill.customer_phone or "")
        ws.cell(row=row, column=5, value=bill.doctor_name or "")
        ws.cell(row=row, column=6, value=bill.payment_method)
        ws.cell(row=row, column=7, value=bill.subtotal)
        ws.cell(row=row, column=8, value=bill.discount_amount)
        ws.cell(row=row, column=9, value=bill.tax_amount)
        ws.cell(row=row, column=10, value=bill.total_amount)
        ws.cell(row=row, column=11, value=bill.amount_paid)
        ws.cell(row=row, column=12, value=bill.change_returned)
        ws.cell(row=row, column=13, value=bill.staff_name)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"bills_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ─── ANALYTICS ────────────────────────────────────────────────────────────────

@router.get("/analytics/overview")
def get_analytics_overview(
    days: int = Query(30, description="Number of days to analyze"),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get comprehensive analytics overview"""
    staff, shop_id = current_user

    cache_key = f"billing_analytics:{shop_id}:{days}"
    cached = dashboard_cache.get(cache_key)
    if cached is not None:
        return cached

    end_date = date.today()
    start_date = end_date - timedelta(days=days)

    records = db.query(DailyRecord).filter(
        DailyRecord.shop_id == shop_id,
        DailyRecord.record_date >= start_date,
        DailyRecord.record_date <= end_date
    ).order_by(DailyRecord.record_date).all()

    if not records:
        return {"message": "No data available", "data": {}}

    total_sales = sum(r.cash_sales + r.online_sales for r in records)
    total_bills = sum(r.no_of_bills for r in records)
    total_expenses = sum(r.total_expenses for r in records)
    avg_daily_sales = total_sales / len(records) if records else 0
    avg_bills_per_day = total_bills / len(records) if records else 0

    daily_data = []
    for r in records:
        daily_data.append({
            "date": str(r.record_date),
            "day": r.record_date.strftime('%A'),
            "sales": float(r.cash_sales + r.online_sales),
            "bills": r.no_of_bills,
            "expenses": float(r.total_expenses),
            "cash_sales": float(r.cash_sales),
            "online_sales": float(r.online_sales),
            "average_bill": float(r.cash_sales + r.online_sales) / r.no_of_bills if r.no_of_bills > 0 else 0
        })

    day_wise = {}
    for r in records:
        day = r.record_date.strftime('%A')
        if day not in day_wise:
            day_wise[day] = {"sales": 0, "bills": 0, "count": 0}
        day_wise[day]["sales"] += float(r.cash_sales + r.online_sales)
        day_wise[day]["bills"] += r.no_of_bills
        day_wise[day]["count"] += 1

    day_wise_avg = {
        day: {
            "avg_sales": data["sales"] / data["count"],
            "avg_bills": data["bills"] / data["count"]
        }
        for day, data in day_wise.items()
    }

    expenses = db.query(
        DailyExpense.expense_category,
        func.sum(DailyExpense.amount).label('total')
    ).join(DailyRecord).filter(
        DailyRecord.shop_id == shop_id,
        DailyRecord.record_date >= start_date,
        DailyRecord.record_date <= end_date
    ).group_by(DailyExpense.expense_category).all()

    expense_breakdown = [{"category": e[0], "amount": float(e[1])} for e in expenses]

    if len(records) >= 7:
        recent_7_days = records[-7:]
        avg_recent = sum(r.cash_sales + r.online_sales for r in recent_7_days) / 7
        prediction_next_7_days = avg_recent * 7
    else:
        prediction_next_7_days = avg_daily_sales * 7

    result = {
        "summary": {
            "period_days": days,
            "total_sales": round(total_sales, 2),
            "total_bills": total_bills,
            "total_expenses": round(total_expenses, 2),
            "net_revenue": round(total_sales - total_expenses, 2),
            "avg_daily_sales": round(avg_daily_sales, 2),
            "avg_bills_per_day": round(avg_bills_per_day, 2),
            "avg_bill_value": round(total_sales / total_bills, 2) if total_bills > 0 else 0
        },
        "daily_trends": daily_data,
        "day_wise_analysis": day_wise_avg,
        "expense_breakdown": expense_breakdown,
        "predictions": {
            "next_7_days_sales": round(prediction_next_7_days, 2),
            "avg_daily_prediction": round(prediction_next_7_days / 7, 2)
        }
    }
    dashboard_cache.set(cache_key, result, ttl=60)
    return result

@router.get("/analytics/comparison")
def get_period_comparison(
    current_days: int = Query(30),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Compare current period with previous period"""
    staff, shop_id = current_user

    cache_key = f"billing_comparison:{shop_id}:{current_days}"
    cached = dashboard_cache.get(cache_key)
    if cached is not None:
        return cached

    end_date = date.today()
    current_start = end_date - timedelta(days=current_days)
    previous_start = current_start - timedelta(days=current_days)

    current_records = db.query(DailyRecord).filter(
        DailyRecord.shop_id == shop_id,
        DailyRecord.record_date >= current_start,
        DailyRecord.record_date <= end_date
    ).all()

    previous_records = db.query(DailyRecord).filter(
        DailyRecord.shop_id == shop_id,
        DailyRecord.record_date >= previous_start,
        DailyRecord.record_date < current_start
    ).all()

    def calculate_metrics(records):
        if not records:
            return {"sales": 0, "bills": 0, "expenses": 0}
        return {
            "sales": sum(r.cash_sales + r.online_sales for r in records),
            "bills": sum(r.no_of_bills for r in records),
            "expenses": sum(r.total_expenses for r in records)
        }

    current = calculate_metrics(current_records)
    previous = calculate_metrics(previous_records)

    def calc_change(curr, prev):
        if prev == 0:
            return 100 if curr > 0 else 0
        return ((curr - prev) / prev) * 100

    result = {
        "current_period": {
            "days": current_days,
            "sales": round(current["sales"], 2),
            "bills": current["bills"],
            "expenses": round(current["expenses"], 2)
        },
        "previous_period": {
            "days": current_days,
            "sales": round(previous["sales"], 2),
            "bills": previous["bills"],
            "expenses": round(previous["expenses"], 2)
        },
        "changes": {
            "sales_change": round(calc_change(current["sales"], previous["sales"]), 2),
            "bills_change": round(calc_change(current["bills"], previous["bills"]), 2),
            "expenses_change": round(calc_change(current["expenses"], previous["expenses"]), 2)
        }
    }
    dashboard_cache.set(cache_key, result, ttl=60)
    return result

@router.get("/profit-analysis")
def get_profit_analysis(
    days: int = Query(30, le=365),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Comprehensive profit & loss analysis for the shop"""
    from collections import defaultdict
    staff, shop_id = current_user

    cache_key = f"billing_profit:{shop_id}:{days}:{start_date}:{end_date}"
    if not search:
        cached = dashboard_cache.get(cache_key)
        if cached is not None:
            return cached

    e_date = end_date or date.today()
    s_date = start_date or (e_date - timedelta(days=days))

    # ── Bills query ──────────────────────────────────────────────────
    bill_q = db.query(models.Bill).filter(
        models.Bill.shop_id == shop_id,
        models.Bill.created_at >= datetime.combine(s_date, datetime.min.time()),
        models.Bill.created_at <= datetime.combine(e_date, datetime.max.time())
    )
    if search:
        bill_q = bill_q.filter(
            (models.Bill.customer_name.ilike(f'%{search}%')) |
            (models.Bill.customer_phone.ilike(f'%{search}%')) |
            (models.Bill.bill_number.ilike(f'%{search}%'))
        )
    bills = bill_q.order_by(models.Bill.created_at.desc()).all()

    # ── Revenue metrics ──────────────────────────────────────────────
    total_revenue = sum(b.total_amount for b in bills)
    total_discounts = sum(b.discount_amount for b in bills)
    total_tax = sum(b.tax_amount for b in bills)
    cash_total = sum(b.cash_amount for b in bills)
    card_total = sum(b.card_amount for b in bills)
    online_total = sum(b.online_amount for b in bills)

    # ── Expenses from daily records ──────────────────────────────────
    records = db.query(DailyRecord).filter(
        DailyRecord.shop_id == shop_id,
        DailyRecord.record_date >= s_date,
        DailyRecord.record_date <= e_date
    ).all()
    total_expenses = sum(r.total_expenses for r in records)
    net_profit = total_revenue - total_expenses
    profit_margin = round(net_profit / total_revenue * 100, 2) if total_revenue > 0 else 0

    # ── Daily P&L trend ──────────────────────────────────────────────
    record_map = {r.record_date: r for r in records}
    daily_revenue = defaultdict(float)
    daily_bills_cnt = defaultdict(int)
    daily_discounts = defaultdict(float)
    for b in bills:
        d = b.created_at.date()
        daily_revenue[d] += b.total_amount
        daily_bills_cnt[d] += 1
        daily_discounts[d] += b.discount_amount

    all_dates = sorted(set(list(daily_revenue.keys()) + list(record_map.keys())))
    daily_pnl = []
    for d in all_dates:
        rev = daily_revenue.get(d, 0)
        exp = float(record_map[d].total_expenses) if d in record_map else 0
        profit = rev - exp
        daily_pnl.append({
            "date": str(d),
            "day": d.strftime('%a'),
            "revenue": round(rev, 2),
            "expenses": round(exp, 2),
            "net_profit": round(profit, 2),
            "profit_margin": round(profit / rev * 100 if rev > 0 else 0, 2),
            "bills": daily_bills_cnt.get(d, 0),
            "discounts": round(daily_discounts.get(d, 0), 2)
        })

    # ── Expense breakdown by category ────────────────────────────────
    expense_cats = db.query(
        DailyExpense.expense_category,
        func.sum(DailyExpense.amount).label('total')
    ).join(DailyRecord).filter(
        DailyRecord.shop_id == shop_id,
        DailyRecord.record_date >= s_date,
        DailyRecord.record_date <= e_date
    ).group_by(DailyExpense.expense_category).all()

    expense_breakdown = [
        {
            "category": e[0],
            "amount": round(float(e[1]), 2),
            "percentage": round(float(e[1]) / total_expenses * 100 if total_expenses > 0 else 0, 2)
        }
        for e in expense_cats
    ]

    # ── Payment split ────────────────────────────────────────────────
    payment_base = (cash_total + card_total + online_total) or total_revenue or 1
    payment_split = {
        "cash": round(cash_total, 2),
        "card": round(card_total, 2),
        "online": round(online_total, 2),
        "cash_pct": round(cash_total / payment_base * 100, 2),
        "card_pct": round(card_total / payment_base * 100, 2),
        "online_pct": round(online_total / payment_base * 100, 2)
    }

    # ── Top items by revenue ─────────────────────────────────────────
    top_items_q = db.query(
        models.BillItem.item_name,
        func.sum(models.BillItem.quantity).label('total_quantity'),
        func.sum(models.BillItem.total_price).label('total_revenue'),
        func.avg(models.BillItem.discount_percent).label('avg_discount'),
        func.count(models.BillItem.id).label('transaction_count')
    ).join(models.Bill).filter(
        models.Bill.shop_id == shop_id,
        models.Bill.created_at >= datetime.combine(s_date, datetime.min.time()),
        models.Bill.created_at <= datetime.combine(e_date, datetime.max.time())
    ).group_by(models.BillItem.item_name).order_by(
        func.sum(models.BillItem.total_price).desc()
    ).limit(10).all()

    top_items = [
        {
            "item_name": r.item_name,
            "total_quantity": r.total_quantity,
            "total_revenue": round(float(r.total_revenue), 2),
            "avg_discount_pct": round(float(r.avg_discount or 0), 2),
            "transaction_count": r.transaction_count
        }
        for r in top_items_q
    ]

    # ── Staff performance ────────────────────────────────────────────
    staff_perf_q = db.query(
        models.Bill.staff_name,
        func.count(models.Bill.id).label('bill_count'),
        func.sum(models.Bill.total_amount).label('total_revenue')
    ).filter(
        models.Bill.shop_id == shop_id,
        models.Bill.created_at >= datetime.combine(s_date, datetime.min.time()),
        models.Bill.created_at <= datetime.combine(e_date, datetime.max.time())
    ).group_by(models.Bill.staff_name).order_by(func.sum(models.Bill.total_amount).desc()).all()

    staff_performance = [
        {
            "staff_name": r.staff_name,
            "bill_count": r.bill_count,
            "total_revenue": round(float(r.total_revenue), 2),
            "avg_bill": round(float(r.total_revenue) / r.bill_count if r.bill_count > 0 else 0, 2)
        }
        for r in staff_perf_q
    ]

    # ── Bill list (for search table) ─────────────────────────────────
    bill_list = [
        {
            "id": b.id,
            "bill_number": b.bill_number,
            "date": b.created_at.strftime('%Y-%m-%d %H:%M'),
            "customer_name": b.customer_name or "Walk-in",
            "customer_phone": b.customer_phone or "",
            "staff_name": b.staff_name,
            "subtotal": round(b.subtotal, 2),
            "discount_amount": round(b.discount_amount, 2),
            "tax_amount": round(b.tax_amount, 2),
            "total_amount": round(b.total_amount, 2),
            "payment_method": b.payment_method,
            "items_count": len(b.items)
        }
        for b in bills[:200]
    ]

    result = {
        "summary": {
            "period_days": days,
            "start_date": str(s_date),
            "end_date": str(e_date),
            "total_revenue": round(total_revenue, 2),
            "total_bills": len(bills),
            "avg_bill_value": round(total_revenue / len(bills) if bills else 0, 2),
            "total_discounts_given": round(total_discounts, 2),
            "discount_rate": round(total_discounts / (total_revenue + total_discounts) * 100 if (total_revenue + total_discounts) > 0 else 0, 2),
            "total_tax_collected": round(total_tax, 2),
            "total_expenses": round(total_expenses, 2),
            "net_profit": round(net_profit, 2),
            "profit_margin": profit_margin
        },
        "daily_pnl": daily_pnl,
        "expense_breakdown": expense_breakdown,
        "payment_split": payment_split,
        "top_items": top_items,
        "staff_performance": staff_performance,
        "bill_list": bill_list
    }

    if not search:
        dashboard_cache.set(cache_key, result, ttl=60)
    return result

# ─── DAILY RECORDS ────────────────────────────────────────────────────────────

@router.get("/daily-records/{record_date}", response_model=daily_records_schemas.DailyRecordResponse)
def get_daily_record(
    record_date: date,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get daily record for specific date"""
    staff, shop_id = current_user

    record = db.query(DailyRecord).filter(
        DailyRecord.shop_id == shop_id,
        DailyRecord.record_date == record_date
    ).first()

    if not record:
        record = DailyRecordsService.update_daily_record(
            db, shop_id, record_date, staff.id, staff.name, {}
        )

    return DailyRecordsService.get_daily_record_with_calculations(db, record)

@router.post("/daily-records", response_model=daily_records_schemas.DailyRecordResponse)
def create_or_update_daily_record(
    data: daily_records_schemas.DailyRecordCreate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Create or update daily record"""
    staff, shop_id = current_user

    record = DailyRecordsService.update_daily_record(
        db,
        shop_id,
        data.record_date,
        staff.id,
        staff.name,
        data.model_dump(exclude={'expenses'})
    )

    for expense_data in data.expenses:
        DailyRecordsService.add_expense(
            db,
            shop_id,
            record.id,
            expense_data.model_dump()
        )

    db.refresh(record)
    return DailyRecordsService.get_daily_record_with_calculations(db, record)

@router.put("/daily-records/{record_date}", response_model=daily_records_schemas.DailyRecordResponse)
def update_daily_record(
    record_date: date,
    data: daily_records_schemas.DailyRecordUpdate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Update daily record"""
    staff, shop_id = current_user

    record = DailyRecordsService.update_daily_record(
        db,
        shop_id,
        record_date,
        staff.id,
        staff.name,
        data.model_dump(exclude_unset=True)
    )

    return DailyRecordsService.get_daily_record_with_calculations(db, record)

@router.post("/daily-records/{record_date}/expenses", response_model=daily_records_schemas.DailyExpense)
def add_expense(
    record_date: date,
    expense: daily_records_schemas.DailyExpenseCreate,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Add expense to daily record"""
    staff, shop_id = current_user

    record = db.query(DailyRecord).filter(
        DailyRecord.shop_id == shop_id,
        DailyRecord.record_date == record_date
    ).first()

    if not record:
        record = DailyRecordsService.get_or_create_daily_record(
            db, shop_id, record_date, staff.id, staff.name
        )

    return DailyRecordsService.add_expense(
        db,
        shop_id,
        record.id,
        expense.model_dump(),
        staff.id,
        staff.name
    )

@router.delete("/daily-records/expenses/{expense_id}")
def delete_expense(
    expense_id: int,
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Delete expense"""
    staff, shop_id = current_user

    expense = db.query(DailyExpense).filter(
        DailyExpense.id == expense_id,
        DailyExpense.shop_id == shop_id
    ).first()

    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")

    daily_record_id = expense.daily_record_id
    db.delete(expense)

    record = db.query(DailyRecord).filter(DailyRecord.id == daily_record_id).first()
    if record:
        total = db.query(func.sum(DailyExpense.amount)).filter(
            DailyExpense.daily_record_id == daily_record_id
        ).scalar() or 0.0
        record.total_expenses = float(total)

    db.commit()
    return {"message": "Expense deleted"}

@router.get("/daily-records", response_model=List[daily_records_schemas.DailyRecordResponse])
def get_daily_records(
    start_date: date = Query(...),
    end_date: date = Query(...),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Get daily records for date range"""
    staff, shop_id = current_user

    records = db.query(DailyRecord).filter(
        DailyRecord.shop_id == shop_id,
        DailyRecord.record_date >= start_date,
        DailyRecord.record_date <= end_date
    ).order_by(DailyRecord.record_date.desc()).all()

    return [DailyRecordsService.get_daily_record_with_calculations(db, r) for r in records]

@router.get("/daily-records/export/excel")
def export_daily_records_excel(
    start_date: date = Query(...),
    end_date: date = Query(...),
    db: Session = Depends(get_db),
    current_user: tuple = Depends(get_current_user)
):
    """Export daily records to Excel"""
    staff, shop_id = current_user

    if start_date > end_date:
        raise HTTPException(status_code=400, detail="Start date cannot be after end date")

    records = db.query(DailyRecord).filter(
        DailyRecord.shop_id == shop_id,
        DailyRecord.record_date >= start_date,
        DailyRecord.record_date <= end_date
    ).order_by(DailyRecord.record_date).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Records"

    headers = [
        "Date", "Day", "No. of Bills", "Average Bill", "Software Sales", "Cash Sales",
        "Depositable Amount", "Small Denomination", "Online Sales", "Total Sales",
        "Unbilled Amount", "Recorded Sales", "Difference", "Cash Deposited",
        "Cash Reserve", "Total Expenses", "Expense Details"
    ]

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, record in enumerate(records, 2):
        calc = DailyRecordsService.get_daily_record_with_calculations(db, record)
        record_date = calc['record_date']
        day_name = record_date.strftime('%A') if hasattr(record_date, 'strftime') else ''

        expense_details = ""
        if calc.get('expenses'):
            expense_lines = []
            for exp in calc['expenses']:
                expense_lines.append(f"{exp.expense_category}: ₹{exp.amount:.2f} ({exp.description or 'N/A'}) - {exp.staff_name or 'Unknown'}")
            expense_details = "; ".join(expense_lines)

        ws.cell(row=row_idx, column=1, value=str(record_date))
        ws.cell(row=row_idx, column=2, value=day_name)
        ws.cell(row=row_idx, column=3, value=calc['no_of_bills'])
        ws.cell(row=row_idx, column=4, value=round(calc['average_bill'], 2))
        ws.cell(row=row_idx, column=5, value=round(calc['software_sales'], 2))
        ws.cell(row=row_idx, column=6, value=round(calc['cash_sales'], 2))
        ws.cell(row=row_idx, column=7, value=round(calc['depositable_amount'], 2))
        ws.cell(row=row_idx, column=8, value=round(calc['small_denomination'], 2))
        ws.cell(row=row_idx, column=9, value=round(calc['online_sales'], 2))
        ws.cell(row=row_idx, column=10, value=round(calc['total_sales'], 2))
        ws.cell(row=row_idx, column=11, value=round(calc['unbilled_amount'], 2))
        ws.cell(row=row_idx, column=12, value=round(calc['recorded_sales'], 2))
        ws.cell(row=row_idx, column=13, value=round(calc['difference'], 2))
        ws.cell(row=row_idx, column=14, value=round(calc['actual_cash_deposited'], 2))
        ws.cell(row=row_idx, column=15, value=round(calc['cash_reserve'], 2))
        ws.cell(row=row_idx, column=16, value=round(calc['total_expenses'], 2))
        ws.cell(row=row_idx, column=17, value=expense_details)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    ws_expenses = wb.create_sheet(title="Expenses")
    expense_headers = ["Date", "Day", "Category", "Amount", "Description", "Staff", "Time"]

    for col, header in enumerate(expense_headers, 1):
        cell = ws_expenses.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    expense_row = 2
    for record in records:
        calc = DailyRecordsService.get_daily_record_with_calculations(db, record)
        record_date = calc['record_date']
        day_name = record_date.strftime('%A') if hasattr(record_date, 'strftime') else ''

        for expense in calc.get('expenses', []):
            ws_expenses.cell(row=expense_row, column=1, value=str(record_date))
            ws_expenses.cell(row=expense_row, column=2, value=day_name)
            ws_expenses.cell(row=expense_row, column=3, value=expense.expense_category)
            ws_expenses.cell(row=expense_row, column=4, value=round(expense.amount, 2))
            ws_expenses.cell(row=expense_row, column=5, value=expense.description or "")
            ws_expenses.cell(row=expense_row, column=6, value=expense.staff_name or "")
            ws_expenses.cell(row=expense_row, column=7, value=str(expense.created_at))
            expense_row += 1

    for col in ws_expenses.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_expenses.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"daily_records_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ─── SHOP CONFIG (staff) ──────────────────────────────────────────────────────

@router.get("/shop/bill-config")
def get_shop_bill_config(
    current_user: tuple = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get bill configuration for current shop"""
    staff, shop_id = current_user

    shop = db.query(Shop).filter(Shop.id == shop_id).first()
    if not shop:
        raise HTTPException(status_code=404, detail="Shop not found")

    if shop.bill_config:
        try:
            config = json.loads(shop.bill_config)
            return {"config": config}
        except Exception:
            pass

    return {"config": _get_default_config()}

def _get_default_config():
    """Default bill configuration"""
    return {
        "storeName": "GENERICART MEDICINE STORE",
        "logo": "",
        "dlNumbers": {
            "dl20": "20-RLF20TR2025000266",
            "dl21": "21-RLF21TR2025000259"
        },
        "flNumber": "",
        "address": {
            "line1": "JOYNAGAR BUS STOP, ROLANDSAY ROAD, AGARTALA",
            "state": "Tripura",
            "pincode": "799001"
        },
        "phone": "6909319003",
        "gstIn": "16GDYPP9241P2Z6"
    }

# ─── BILL CONFIG (staff-editable) ─────────────────────────────────────────────

class BillConfigUpdate(BaseModel):
    storeName: str
    logo: str
    dlNumbers: dict
    flNumber: str
    address: dict
    phone: str
    gstIn: str

@router.put("/admin/bill-config")
def update_shop_bill_config(
    config: BillConfigUpdate,
    current_user: tuple = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Staff updates bill configuration for their shop"""
    staff, shop_id = current_user

    shop = db.query(Shop).filter(Shop.id == shop_id).first()
    if not shop:
        raise HTTPException(status_code=404, detail="Shop not found")

    shop.bill_config = json.dumps(config.dict())
    db.commit()

    return {"message": "Bill configuration updated successfully"}

@router.get("/admin/bill-config")
def get_shop_bill_config_admin(
    current_user: tuple = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Staff gets bill configuration for their shop"""
    staff, shop_id = current_user

    shop = db.query(Shop).filter(Shop.id == shop_id).first()
    if not shop:
        raise HTTPException(status_code=404, detail="Shop not found")

    if shop.bill_config:
        try:
            config = json.loads(shop.bill_config)
            return {"config": config}
        except Exception:
            pass

    return {"config": None}
