from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database.database import get_db
from modules.auth.models import Admin, Shop
from typing import Optional
from datetime import datetime, date, timedelta
from io import BytesIO
import json, math
from collections import defaultdict
from modules.billing_v2.daily_records_models import DailyRecord, DailyExpense
from pydantic import BaseModel
from modules.auth.dependencies import get_current_admin
from modules.billing_v2 import models
from modules.billing_v2.admin.admin_analytics_service import BillingAdminAnalytics
from app.utils.cache import dashboard_cache
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()

# ─── ANALYTICS ────────────────────────────────────────────────────────────────

@router.get("/admin/analytics/dashboard")
def get_admin_dashboard(
    shop_id: Optional[int] = None,
    days: int = Query(30, le=365),
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    """Get comprehensive billing analytics dashboard for admin"""
    cache_key = f"billing_admin_dashboard:{admin.organization_id}:{shop_id or 'all'}:{days}"
    cached = dashboard_cache.get(cache_key)
    if cached is not None:
        return cached

    analytics = BillingAdminAnalytics()
    data = analytics._gather_billing_data(db, admin.organization_id, shop_id, days)
    dashboard_cache.set(cache_key, data, ttl=60)
    return data

@router.get("/admin/analytics/ai-insights")
def get_ai_insights(
    shop_id: Optional[int] = None,
    days: int = Query(30, le=365),
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    """Generate AI-powered insights for billing data"""
    cache_key = f"billing_admin_ai:{admin.organization_id}:{shop_id or 'all'}:{days}"
    cached = dashboard_cache.get(cache_key)
    if cached is not None:
        return cached

    analytics = BillingAdminAnalytics()
    result = analytics.generate_comprehensive_analysis(
        db, admin.organization_id, shop_id, days
    )
    dashboard_cache.set(cache_key, result, ttl=3600)
    return result

# ─── BILL MANAGEMENT ──────────────────────────────────────────────────────────

@router.get("/admin/bills")
def get_admin_bills(
    shop_id: Optional[int] = Query(None),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    customer_phone: Optional[str] = None,
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    """List bills for admin (org-scoped, paginated, with optional shop filter)"""
    query = (
        db.query(models.Bill)
        .join(Shop, models.Bill.shop_id == Shop.id)
        .filter(Shop.organization_id == admin.organization_id)
    )
    if shop_id:
        query = query.filter(models.Bill.shop_id == shop_id)
    if start_date:
        query = query.filter(models.Bill.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.filter(models.Bill.created_at <= datetime.combine(end_date, datetime.max.time()))
    if customer_phone:
        query = query.filter(models.Bill.customer_phone == customer_phone)

    total = query.count()
    bills = query.order_by(models.Bill.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

    return {
        "items": bills,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": math.ceil(total / per_page) if total > 0 else 1
    }

@router.get("/admin/bills/{bill_id}")
def get_admin_bill(
    bill_id: int,
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    """Get a single bill by ID (org-scoped)"""
    bill = (
        db.query(models.Bill)
        .join(Shop, models.Bill.shop_id == Shop.id)
        .filter(models.Bill.id == bill_id, Shop.organization_id == admin.organization_id)
        .first()
    )
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    return bill

@router.get("/admin/shop/{shop_id}/bill-config")
def get_admin_shop_bill_config(
    shop_id: int,
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    """Get bill config for a specific shop (for printing bills in admin view)"""
    shop = db.query(Shop).filter(
        Shop.id == shop_id,
        Shop.organization_id == admin.organization_id
    ).first()
    if not shop:
        raise HTTPException(status_code=404, detail="Shop not found")
    if shop.bill_config:
        try:
            return {"config": json.loads(shop.bill_config)}
        except Exception:
            pass
    return {"config": None}

class AdminBillConfigUpdate(BaseModel):
    storeName: str
    logo: str
    dlNumbers: dict
    flNumber: str
    address: dict
    phone: str
    gstIn: str

@router.put("/admin/shop/{shop_id}/bill-config")
def update_admin_shop_bill_config(
    shop_id: int,
    config: AdminBillConfigUpdate,
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    """Admin updates bill config for a specific shop (org-scoped)"""
    shop = db.query(Shop).filter(
        Shop.id == shop_id,
        Shop.organization_id == admin.organization_id
    ).first()
    if not shop:
        raise HTTPException(status_code=404, detail="Shop not found")
    shop.bill_config = json.dumps(config.dict())
    db.commit()
    return {"message": "Bill configuration updated successfully"}

# ─── REPORTS ──────────────────────────────────────────────────────────────────

@router.get("/admin/top-selling")
def get_admin_top_selling(
    shop_id: Optional[int] = Query(None),
    limit: int = Query(10, le=50),
    days: int = Query(30, le=365),
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    """Top selling items across org (optional shop filter)"""
    cutoff_date = datetime.now() - timedelta(days=days)
    query = (
        db.query(
            models.BillItem.item_name,
            func.sum(models.BillItem.quantity).label('total_quantity'),
            func.sum(models.BillItem.total_price).label('total_revenue'),
            func.count(models.BillItem.id).label('transaction_count')
        )
        .join(models.Bill, models.BillItem.bill_id == models.Bill.id)
        .join(Shop, models.Bill.shop_id == Shop.id)
        .filter(Shop.organization_id == admin.organization_id, models.Bill.created_at >= cutoff_date)
    )
    if shop_id:
        query = query.filter(models.Bill.shop_id == shop_id)

    results = (
        query.group_by(models.BillItem.item_name)
        .order_by(func.sum(models.BillItem.quantity).desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "item_name": r.item_name,
            "total_quantity": r.total_quantity,
            "total_revenue": float(r.total_revenue),
            "transaction_count": r.transaction_count
        }
        for r in results
    ]

@router.get("/admin/daily-sales")
def get_admin_daily_sales(
    shop_id: Optional[int] = Query(None),
    days: int = Query(7, le=90),
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    """Daily sales across org (optional shop filter)"""
    cutoff_date = datetime.now() - timedelta(days=days)
    query = (
        db.query(
            func.date(models.Bill.created_at).label('date'),
            func.count(models.Bill.id).label('bill_count'),
            func.sum(models.Bill.total_amount).label('total_sales')
        )
        .join(Shop, models.Bill.shop_id == Shop.id)
        .filter(Shop.organization_id == admin.organization_id, models.Bill.created_at >= cutoff_date)
    )
    if shop_id:
        query = query.filter(models.Bill.shop_id == shop_id)

    results = (
        query.group_by(func.date(models.Bill.created_at))
        .order_by(func.date(models.Bill.created_at).desc())
        .all()
    )
    return [
        {"date": str(r.date), "bill_count": r.bill_count, "total_sales": float(r.total_sales)}
        for r in results
    ]

@router.get("/admin/export/bills")
def export_admin_bills_excel(
    shop_id: Optional[int] = Query(None),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    """Export bills to Excel (org-scoped, optional shop filter)"""
    query = (
        db.query(models.Bill)
        .join(Shop, models.Bill.shop_id == Shop.id)
        .filter(Shop.organization_id == admin.organization_id)
    )
    if shop_id:
        query = query.filter(models.Bill.shop_id == shop_id)
    if start_date:
        query = query.filter(models.Bill.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.filter(models.Bill.created_at <= datetime.combine(end_date, datetime.max.time()))

    bills = query.order_by(models.Bill.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bills"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Bill Number", "Date", "Shop", "Customer Name", "Customer Phone",
               "Doctor Name", "Payment Method", "Subtotal", "Discount", "Tax",
               "Total Amount", "Amount Paid", "Change", "Staff Name"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Build shop_id → shop_name map for display
    shop_ids = list({b.shop_id for b in bills})
    shops = db.query(Shop).filter(Shop.id.in_(shop_ids)).all()
    shop_name_map = {s.id: s.shop_name for s in shops}

    for row, bill in enumerate(bills, 2):
        ws.cell(row=row, column=1, value=bill.bill_number)
        ws.cell(row=row, column=2, value=bill.created_at.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=3, value=shop_name_map.get(bill.shop_id, str(bill.shop_id)))
        ws.cell(row=row, column=4, value=bill.customer_name or "")
        ws.cell(row=row, column=5, value=bill.customer_phone or "")
        ws.cell(row=row, column=6, value=bill.doctor_name or "")
        ws.cell(row=row, column=7, value=bill.payment_method)
        ws.cell(row=row, column=8, value=bill.subtotal)
        ws.cell(row=row, column=9, value=bill.discount_amount)
        ws.cell(row=row, column=10, value=bill.tax_amount)
        ws.cell(row=row, column=11, value=bill.total_amount)
        ws.cell(row=row, column=12, value=bill.amount_paid)
        ws.cell(row=row, column=13, value=bill.change_returned)
        ws.cell(row=row, column=14, value=bill.staff_name)

    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"admin_bills_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ─── PROFIT ANALYSIS ──────────────────────────────────────────────────────────

@router.get("/admin/profit-analysis")
def get_admin_profit_analysis(
    shop_id: Optional[int] = Query(None),
    days: int = Query(30, le=365),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    """Comprehensive profit & loss analysis for admin (org-scoped, optional shop filter)"""
    cache_key = f"billing_admin_profit:{admin.organization_id}:{shop_id or 'all'}:{days}:{start_date}:{end_date}"
    if not search:
        cached = dashboard_cache.get(cache_key)
        if cached is not None:
            return cached

    e_date = end_date or date.today()
    s_date = start_date or (e_date - timedelta(days=days))

    # ── Base bill query (org-scoped) ─────────────────────────────────
    bill_q = (
        db.query(models.Bill)
        .join(Shop, models.Bill.shop_id == Shop.id)
        .filter(
            Shop.organization_id == admin.organization_id,
            models.Bill.created_at >= datetime.combine(s_date, datetime.min.time()),
            models.Bill.created_at <= datetime.combine(e_date, datetime.max.time())
        )
    )
    if shop_id:
        bill_q = bill_q.filter(models.Bill.shop_id == shop_id)
    if search:
        bill_q = bill_q.filter(
            (models.Bill.customer_name.ilike(f'%{search}%')) |
            (models.Bill.customer_phone.ilike(f'%{search}%')) |
            (models.Bill.bill_number.ilike(f'%{search}%'))
        )
    bills = bill_q.order_by(models.Bill.created_at.desc()).all()

    # ── Revenue metrics ──────────────────────────────────────────────
    total_revenue = sum(b.total_amount for b in bills)
    total_discounts = sum(b.discount_amount for b in bills)
    total_tax = sum(b.tax_amount for b in bills)
    cash_total = sum(b.cash_amount for b in bills)
    card_total = sum(b.card_amount for b in bills)
    online_total = sum(b.online_amount for b in bills)

    # ── Expenses ─────────────────────────────────────────────────────
    record_q = (
        db.query(DailyRecord)
        .join(Shop, DailyRecord.shop_id == Shop.id)
        .filter(
            Shop.organization_id == admin.organization_id,
            DailyRecord.record_date >= s_date,
            DailyRecord.record_date <= e_date
        )
    )
    if shop_id:
        record_q = record_q.filter(DailyRecord.shop_id == shop_id)
    records = record_q.all()
    total_expenses = sum(r.total_expenses for r in records)
    net_profit = total_revenue - total_expenses
    profit_margin = round(net_profit / total_revenue * 100, 2) if total_revenue > 0 else 0

    # ── Daily P&L trend ──────────────────────────────────────────────
    record_map = defaultdict(float)
    for r in records:
        record_map[r.record_date] += r.total_expenses

    daily_revenue = defaultdict(float)
    daily_bills_cnt = defaultdict(int)
    daily_discounts_map = defaultdict(float)
    for b in bills:
        d = b.created_at.date()
        daily_revenue[d] += b.total_amount
        daily_bills_cnt[d] += 1
        daily_discounts_map[d] += b.discount_amount

    all_dates = sorted(set(list(daily_revenue.keys()) + list(record_map.keys())))
    daily_pnl = []
    for d in all_dates:
        rev = daily_revenue.get(d, 0)
        exp = record_map.get(d, 0)
        profit = rev - exp
        daily_pnl.append({
            "date": str(d),
            "day": d.strftime('%a'),
            "revenue": round(rev, 2),
            "expenses": round(exp, 2),
            "net_profit": round(profit, 2),
            "profit_margin": round(profit / rev * 100 if rev > 0 else 0, 2),
            "bills": daily_bills_cnt.get(d, 0),
            "discounts": round(daily_discounts_map.get(d, 0), 2)
        })

    # ── Expense breakdown ────────────────────────────────────────────
    exp_q = (
        db.query(DailyExpense.expense_category, func.sum(DailyExpense.amount).label('total'))
        .join(DailyRecord)
        .join(Shop, DailyRecord.shop_id == Shop.id)
        .filter(
            Shop.organization_id == admin.organization_id,
            DailyRecord.record_date >= s_date,
            DailyRecord.record_date <= e_date
        )
    )
    if shop_id:
        exp_q = exp_q.filter(DailyRecord.shop_id == shop_id)
    expense_cats = exp_q.group_by(DailyExpense.expense_category).all()

    expense_breakdown = [
        {
            "category": e[0],
            "amount": round(float(e[1]), 2),
            "percentage": round(float(e[1]) / total_expenses * 100 if total_expenses > 0 else 0, 2)
        }
        for e in expense_cats
    ]

    # ── Payment split ────────────────────────────────────────────────
    payment_base = (cash_total + card_total + online_total) or total_revenue or 1
    payment_split = {
        "cash": round(cash_total, 2),
        "card": round(card_total, 2),
        "online": round(online_total, 2),
        "cash_pct": round(cash_total / payment_base * 100, 2),
        "card_pct": round(card_total / payment_base * 100, 2),
        "online_pct": round(online_total / payment_base * 100, 2)
    }

    # ── Top items by revenue ─────────────────────────────────────────
    top_items_q = (
        db.query(
            models.BillItem.item_name,
            func.sum(models.BillItem.quantity).label('total_quantity'),
            func.sum(models.BillItem.total_price).label('total_revenue'),
            func.avg(models.BillItem.discount_percent).label('avg_discount'),
            func.count(models.BillItem.id).label('transaction_count')
        )
        .join(models.Bill)
        .join(Shop, models.Bill.shop_id == Shop.id)
        .filter(
            Shop.organization_id == admin.organization_id,
            models.Bill.created_at >= datetime.combine(s_date, datetime.min.time()),
            models.Bill.created_at <= datetime.combine(e_date, datetime.max.time())
        )
    )
    if shop_id:
        top_items_q = top_items_q.filter(models.Bill.shop_id == shop_id)
    top_items_res = (
        top_items_q.group_by(models.BillItem.item_name)
        .order_by(func.sum(models.BillItem.total_price).desc())
        .limit(10).all()
    )
    top_items = [
        {
            "item_name": r.item_name,
            "total_quantity": r.total_quantity,
            "total_revenue": round(float(r.total_revenue), 2),
            "avg_discount_pct": round(float(r.avg_discount or 0), 2),
            "transaction_count": r.transaction_count
        }
        for r in top_items_res
    ]

    # ── Staff performance ────────────────────────────────────────────
    staff_q = (
        db.query(
            models.Bill.staff_name,
            func.count(models.Bill.id).label('bill_count'),
            func.sum(models.Bill.total_amount).label('total_revenue')
        )
        .join(Shop, models.Bill.shop_id == Shop.id)
        .filter(
            Shop.organization_id == admin.organization_id,
            models.Bill.created_at >= datetime.combine(s_date, datetime.min.time()),
            models.Bill.created_at <= datetime.combine(e_date, datetime.max.time())
        )
    )
    if shop_id:
        staff_q = staff_q.filter(models.Bill.shop_id == shop_id)
    staff_res = staff_q.group_by(models.Bill.staff_name).order_by(func.sum(models.Bill.total_amount).desc()).all()
    staff_performance = [
        {
            "staff_name": r.staff_name,
            "bill_count": r.bill_count,
            "total_revenue": round(float(r.total_revenue), 2),
            "avg_bill": round(float(r.total_revenue) / r.bill_count if r.bill_count > 0 else 0, 2)
        }
        for r in staff_res
    ]

    # ── Shop comparison (admin only) ─────────────────────────────────
    shop_comparison = []
    if not shop_id:
        shop_rev_q = (
            db.query(
                models.Bill.shop_id,
                Shop.shop_name,
                func.sum(models.Bill.total_amount).label('revenue'),
                func.count(models.Bill.id).label('bills')
            )
            .join(Shop, models.Bill.shop_id == Shop.id)
            .filter(
                Shop.organization_id == admin.organization_id,
                models.Bill.created_at >= datetime.combine(s_date, datetime.min.time()),
                models.Bill.created_at <= datetime.combine(e_date, datetime.max.time())
            )
            .group_by(models.Bill.shop_id, Shop.shop_name)
            .order_by(func.sum(models.Bill.total_amount).desc())
            .all()
        )
        # Get expenses per shop
        shop_exp_q = (
            db.query(DailyRecord.shop_id, func.sum(DailyRecord.total_expenses).label('expenses'))
            .join(Shop, DailyRecord.shop_id == Shop.id)
            .filter(
                Shop.organization_id == admin.organization_id,
                DailyRecord.record_date >= s_date,
                DailyRecord.record_date <= e_date
            )
            .group_by(DailyRecord.shop_id).all()
        )
        shop_exp_map = {r.shop_id: float(r.expenses) for r in shop_exp_q}

        for r in shop_rev_q:
            rev = float(r.revenue)
            exp = shop_exp_map.get(r.shop_id, 0)
            profit = rev - exp
            shop_comparison.append({
                "shop_id": r.shop_id,
                "shop_name": r.shop_name,
                "revenue": round(rev, 2),
                "expenses": round(exp, 2),
                "net_profit": round(profit, 2),
                "profit_margin": round(profit / rev * 100 if rev > 0 else 0, 2),
                "bills": r.bills
            })

    # ── Bill list ────────────────────────────────────────────────────
    bill_list = [
        {
            "id": b.id,
            "bill_number": b.bill_number,
            "date": b.created_at.strftime('%Y-%m-%d %H:%M'),
            "customer_name": b.customer_name or "Walk-in",
            "customer_phone": b.customer_phone or "",
            "staff_name": b.staff_name,
            "subtotal": round(b.subtotal, 2),
            "discount_amount": round(b.discount_amount, 2),
            "tax_amount": round(b.tax_amount, 2),
            "total_amount": round(b.total_amount, 2),
            "payment_method": b.payment_method,
            "items_count": len(b.items)
        }
        for b in bills[:200]
    ]

    result = {
        "summary": {
            "period_days": days,
            "start_date": str(s_date),
            "end_date": str(e_date),
            "total_revenue": round(total_revenue, 2),
            "total_bills": len(bills),
            "avg_bill_value": round(total_revenue / len(bills) if bills else 0, 2),
            "total_discounts_given": round(total_discounts, 2),
            "discount_rate": round(total_discounts / (total_revenue + total_discounts) * 100 if (total_revenue + total_discounts) > 0 else 0, 2),
            "total_tax_collected": round(total_tax, 2),
            "total_expenses": round(total_expenses, 2),
            "net_profit": round(net_profit, 2),
            "profit_margin": profit_margin
        },
        "daily_pnl": daily_pnl,
        "expense_breakdown": expense_breakdown,
        "payment_split": payment_split,
        "top_items": top_items,
        "staff_performance": staff_performance,
        "shop_comparison": shop_comparison,
        "bill_list": bill_list
    }

    if not search:
        dashboard_cache.set(cache_key, result, ttl=60)
    return result
